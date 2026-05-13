# -*- coding: utf-8 -*-
"""
Generate Test Case Excel: UC56-57_66_68 Tin tuc
Date: 2026-05-11
Author: QC Agent
Version: v1
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = r'c:\Users\admin\Desktop\CMC\mbfs-bqa\.claude\skills\qc-tc-design-MOBILE\templates\[MBFS] Template TestCase - Mobile.xlsx'
OUTPUT_PATH = r'c:\Users\admin\Desktop\CMC\mbfs-bqa\docs\QC-MOBILE\testcases\UC56-57_66_68\UC56-57_66_68_tin-tuc_testcases_20260511_v1.xlsx'

# Load template
wb = openpyxl.load_workbook(TEMPLATE_PATH)
ws = wb['Tên tính năng']
ws.title = 'Tin tức'

# Update metadata
ws['B1'] = 'Cổng Một Cửa Đầu Tư Quốc Gia - Ứng dụng di động'
ws['D1'] = 'MBFS_UC56-57_66_68_TinTuc_Testcase_v1.0'
ws['D3'] = 'Khai thác tin tức công bố trên Mobile'

# Clear data from Row 8 onwards
if ws.max_row >= 8:
    ws.delete_rows(8, ws.max_row - 7)

# Define styles
thin_side = Side(style='thin')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
section_fill = PatternFill(start_color='FFA4C2F4', end_color='FFA4C2F4', fill_type='solid')
check_fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
sub_fill = PatternFill(start_color='FFE3FAFD', end_color='FFE3FAFD', fill_type='solid')
bold_font = Font(name='Times New Roman', size=11, bold=True)
tc_font = Font(name='Times New Roman', size=11)
tc_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Column widths
col_widths = {'A': 14.25, 'B': 39.0, 'C': 37.38, 'D': 50.38, 'E': 48.63, 'F': 16.0, 'G': 16.0, 'K': 16.0, 'O': 16.0, 'Q': 16.0}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# ===== DATA LIST =====
# Format: ("section", text), ("check", text), ("sub", text),
#         ("tc", id, title, pre, step, expected, note)
data = [
    # ============ SECTION 1: Man hinh Danh sach Tin tuc ============
    ("section", "1. Màn hình Danh sách Tin tức"),
    ("check", "Check UI/UX"),
    ("tc", "TC_001", "Kiểm tra UI/UX màn hình Danh sách Tin tức khi có dữ liệu",
     "1. Đăng nhập thành công vào ứng dụng\n2. Có dữ liệu bài viết trong hệ thống",
     "1. Từ Sidebar, nhấp vào mục \"Tin tức\"\n2. Kiểm tra hiển thị màn hình [Danh sách Tin tức]",
     "2.\n- Hiển thị đầy đủ các item, màu sắc, font chữ, layout giống Design\n(Tham khảo ảnh Màn hình Danh sách Tin tức sheet WFDesign)", ""),
    ("tc", "TC_002", "Kiểm tra Header hiển thị tiêu đề \"Tin tức\" cố định",
     "1. Đăng nhập thành công\n2. Mở màn hình [Danh sách Tin tức]",
     "1. Kiểm tra hiển thị Header màn hình [Danh sách Tin tức]",
     "1.\n- Header nền đỏ đậm\n- Tiêu đề \"Tin tức\" nằm giữa, màu trắng\n- Icon kính lúp (tìm kiếm) màu trắng bên phải\n- Icon filter (slider) màu trắng ngoài cùng bên phải", ""),
    ("tc", "TC_003", "Kiểm tra Tab Bar hiển thị đầy đủ 19 Tab",
     "1. Đăng nhập thành công\n2. Mở màn hình [Danh sách Tin tức]",
     "1. Kiểm tra hiển thị Tab Bar bên dưới Header",
     "1.\n- Hiển thị đầy đủ 19 Tab: Tất cả, Chính sách nổi bật, Chính sách đầu tư, Dịch vụ công, Kinh tế, Văn hóa, Giao thông, Y tế, Lao động, Xã hội, Du lịch, Thể thao, Quốc tế, Chính trị, Thời sự, Giáo dục, Tài chính, Câu chuyện thành công, Khác\n- Tab \"Tất cả\" được chọn mặc định (nền đỏ đậm, chữ trắng)\n- Các tab còn lại: nền xám nhạt, chữ đen/xanh đen\n- Hỗ trợ cuộn ngang", ""),
    ("tc", "TC_004", "Kiểm tra section \"Tin nổi bật\" khi có ≥2 bài",
     "1. Đăng nhập thành công\n2. Mở màn hình [Danh sách Tin tức]\n3. Có ≥2 bài viết được đánh dấu \"Tin nổi bật\"",
     "1. Kiểm tra hiển thị section \"Tin nổi bật\"",
     "1.\n- Hiển thị carousel cuộn ngang\n- Tự động chuyển bài sau mỗi 5 giây\n- Không có dot indicator\n- Mỗi card: nền ảnh cover + overlay tối, Tag đỏ \"Tin nổi bật\" góc trên trái, Tiêu đề trắng đậm (max 2 dòng), Trích dẫn (max 2 dòng), Footer: icon đồng hồ + ngày DD/MM/YYYY + icon user + tên tác giả", ""),
    ("tc", "TC_005", "Kiểm tra section \"Tin nổi bật\" khi chỉ có 1 bài",
     "1. Đăng nhập thành công\n2. Mở màn hình [Danh sách Tin tức]\n3. Chỉ có 1 bài viết được đánh dấu \"Tin nổi bật\"",
     "1. Kiểm tra hiển thị section \"Tin nổi bật\"",
     "1.\n- Hiển thị 1 card tĩnh, không cuộn ngang\n- Không có auto-scroll\n- Card hiển thị đầy đủ: Tag đỏ, Tiêu đề, Trích dẫn, Footer", ""),
    ("tc", "TC_006", "Kiểm tra section \"Tin nổi bật\" khi không có bài nào",
     "1. Đăng nhập thành công\n2. Mở màn hình [Danh sách Tin tức]\n3. Không có bài viết nào được đánh dấu \"Tin nổi bật\"",
     "1. Kiểm tra hiển thị section \"Tin nổi bật\"",
     "1.\n- Section \"Tin nổi bật\" ẩn hoàn toàn\n- Không chiếm không gian trên màn hình\n- Section \"Tin tức mới nhất\" được kéo lên trên", ""),
    ("tc", "TC_007", "Kiểm tra section \"Tin tức mới nhất\" hiển thị đúng cấu trúc",
     "1. Đăng nhập thành công\n2. Mở màn hình [Danh sách Tin tức]\n3. Có dữ liệu bài viết",
     "1. Kiểm tra hiển thị section \"Tin tức mới nhất\"",
     "1.\n- Label \"Tin tức mới nhất\" font đậm, màu đen/xanh đen, căn trái\n- Danh sách cuộn dọc\n- Mỗi card: Thumbnail bên trái (hình vuông, bo góc), bên phải gồm Tag Category (nền nhạt, chữ đỏ), Tiêu đề đậm đen (max 2 dòng), Trích dẫn xám (max 2 dòng), Footer: icon đồng hồ + ngày + icon user + tên tác giả", ""),
    # PLACEHOLDER_TC008_ONWARDS
]