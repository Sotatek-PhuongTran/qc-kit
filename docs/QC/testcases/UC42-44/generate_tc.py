import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

template_path = r"d:/mbfs/mbfs-bqa/.claude/skills/qc-tc-design-MOBILE/templates/[MBFS] Template TestCase - Mobile.xlsx"
output_path = r"d:/mbfs/mbfs-bqa/docs/QC-MOBILE/testcases/UC42-44/UC42-44_quan-ly-dat-lich_testcases_20260508_v1.xlsx"

shutil.copy2(template_path, output_path)
wb = load_workbook(output_path)

ws = wb["Testcase"]
ws.title = "Quan ly dat lich"

font_default = Font(name='Calibri', size=11, color='FF000000')
font_bold = Font(name='Calibri', size=11, color='FF000000', bold=True)
align_lw = Alignment(horizontal='left', vertical='center', wrap_text=True)
fill_section = PatternFill(start_color='FFA4C2F4', end_color='FFA4C2F4', fill_type='solid')
fill_check = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
fill_sub = PatternFill(start_color='FFE3FAFD', end_color='FFE3FAFD', fill_type='solid')

ws.cell(row=1, column=2).value = "Cổng Một Cửa Đầu Tư Quốc Gia - Ứng dụng di động"
ws.cell(row=1, column=4).value = "MBFS_UC-42-44_QuanLyDatLich_Testcase_v1.0"
ws.cell(row=3, column=4).value = "Quản lý đặt lịch nộp hồ sơ (Mobile)"
ws.cell(row=4, column=2).value = "QC Agent"
ws.cell(row=4, column=4).value = "2026-05-08"

for merge in list(ws.merged_cells.ranges):
    if merge.min_row >= 8:
        ws.unmerge_cells(str(merge))
for row in range(8, ws.max_row + 1):
    for col in range(1, 19):
        ws.cell(row=row, column=col).value = None
        ws.cell(row=row, column=col).fill = PatternFill()

cur = [8]
tc_counter = [0]

def section(title):
    r = cur[0]
    for c in range(1, 19):
        ws.cell(row=r, column=c).fill = fill_section
    ws.cell(row=r, column=2).value = title
    ws.cell(row=r, column=2).font = font_bold
    ws.cell(row=r, column=2).alignment = align_lw
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
    cur[0] += 1

def check_type(title):
    r = cur[0]
    for c in range(1, 19):
        ws.cell(row=r, column=c).fill = fill_check
    ws.cell(row=r, column=2).value = title
    ws.cell(row=r, column=2).font = font_bold
    ws.cell(row=r, column=2).alignment = align_lw
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
    cur[0] += 1

def sub_label(title):
    r = cur[0]
    for c in range(1, 19):
        ws.cell(row=r, column=c).fill = fill_sub
    ws.cell(row=r, column=2).value = title
    ws.cell(row=r, column=2).font = font_bold
    ws.cell(row=r, column=2).alignment = align_lw
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
    cur[0] += 1

def tc(title, pre, steps, expected, note=""):
    r = cur[0]
    tc_counter[0] += 1
    ws.cell(row=r, column=1).value = f"TC_{tc_counter[0]:03d}"
    ws.cell(row=r, column=1).font = font_default
    ws.cell(row=r, column=1).alignment = align_lw
    ws.cell(row=r, column=2).value = title
    ws.cell(row=r, column=2).font = font_default
    ws.cell(row=r, column=2).alignment = align_lw
    ws.cell(row=r, column=3).value = pre
    ws.cell(row=r, column=3).font = font_default
    ws.cell(row=r, column=3).alignment = align_lw
    ws.cell(row=r, column=4).value = steps
    ws.cell(row=r, column=4).font = font_default
    ws.cell(row=r, column=4).alignment = align_lw
    ws.cell(row=r, column=5).value = expected
    ws.cell(row=r, column=5).font = font_default
    ws.cell(row=r, column=5).alignment = align_lw
    ws.cell(row=r, column=6).value = note
    ws.cell(row=r, column=6).font = font_default
    ws.cell(row=r, column=6).alignment = align_lw
    ws.cell(row=r, column=7).value = "Untested"
    ws.cell(row=r, column=11).value = "Untested"
    cur[0] += 1

# ============================================================
# SECTION 1: Màn hình Danh sách Đặt lịch nộp hồ sơ
# ============================================================
section("1. Màn hình Danh sách Đặt lịch nộp hồ sơ")

# --- Check UI/UX ---
check_type("Check UI/UX")

tc("[Màn hình Danh sách lịch hẹn] Kiểm tra UI/UX khi có dữ liệu",
   "Đăng nhập thành công tài khoản Cá nhân/Tổ chức có ≥1 lịch hẹn.",
   "1. Vào Trang chủ → nhấp Quick Access \"Quản lý đặt lịch\".\n2. Kiểm tra hiển thị Màn hình [Danh sách lịch hẹn].",
   "2.\n- Hiển thị đầy đủ các item, màu sắc, font chữ, layout giống Design\n(Tham khảo ảnh \"UC 42-44_ Tab List\" sheet WFDesign)\n- Header đỏ: nút Back (←) trái, tiêu đề \"Quản lý đặt lịch nộp hồ sơ\" trắng căn giữa.\n- Khung Tìm kiếm + icon [Filter] bên phải.\n- Tab bar 6 tab theo thứ tự: \"Tất cả\" (active đỏ + underline), \"Chờ xác nhận\", \"Đã xác nhận\", \"Đã hủy\", \"Đã bỏ lượt\", \"Đã hoàn thành\".\n- Danh sách Card lịch hẹn.\n- Bottom Navigation.")

tc("[Màn hình Danh sách lịch hẹn] Kiểm tra UI/UX khi rỗng (tài khoản chưa có lịch hẹn nào)",
   "Đăng nhập tài khoản chưa có lịch hẹn nào.",
   "1. Mở Màn hình [Danh sách lịch hẹn].\n2. Kiểm tra hiển thị vùng danh sách tại Tab \"Tất cả\".",
   "2.\n- Header, Search box, Filter icon, Tab bar: hiển thị bình thường.\n- Vùng danh sách: Empty state — icon + text \"Không có dữ liệu.\" (CMR-14) căn giữa.")

tc("[Màn hình Danh sách lịch hẹn] Kiểm tra Loading state toàn màn hình khi first-load",
   "Đăng nhập thành công, đang chờ API trả danh sách (first-load).",
   "1. Tap Quick Access \"Quản lý đặt lịch\".\n2. Kiểm tra hiển thị màn hình trong khi chờ API.",
   "2.\n- Full-screen loading overlay hiển thị trên toàn vùng nội dung (CMR-07, AC17).\n- Sau khi API trả → ẩn overlay, hiển thị danh sách + tab + search.")

tc("[Ô Search] Kiểm tra hiển thị mặc định (icon + placeholder)",
   "Đang ở Màn hình [Danh sách lịch hẹn].",
   "1. Kiểm tra hiển thị ô [Search].",
   "1.\n- Icon kính lúp bên trái trong ô.\n- Placeholder \"Tìm kiếm thủ tục...\".\n- Ô rỗng mặc định.")

tc("[Icon Filter] Kiểm tra hiển thị mặc định (không có indicator)",
   "Đang ở Màn hình [Danh sách lịch hẹn], chưa áp dụng filter.",
   "1. Kiểm tra hiển thị icon [Filter] bên phải ô Search.",
   "1.\n- Icon filter viền bo tròn.\n- Không có chấm indicator màu xanh lá ở góc phải trên.")

tc("[Icon Filter] Kiểm tra hiển thị khi có filter đang active",
   "Đã áp dụng filter (Lĩnh vực hoặc Dịch vụ công ≠ \"Tất cả\").",
   "1. Kiểm tra hiển thị icon [Filter].",
   "1. Hiển thị chấm indicator màu xanh lá cây ở góc phải bên trên của icon filter (CMR-02, AC8).")

tc("[Tab bar] Kiểm tra hiển thị đủ 6 tab đúng thứ tự",
   "Đang ở Màn hình [Danh sách lịch hẹn].",
   "1. Kiểm tra hiển thị Tab bar.",
   "1. Hiển thị đúng 6 tab theo thứ tự từ trái sang phải: \"Tất cả\" → \"Chờ xác nhận\" → \"Đã xác nhận\" → \"Đã hủy\" → \"Đã bỏ lượt\" → \"Đã hoàn thành\" (AC1).")

tc("[Tab bar] Kiểm tra tab mặc định \"Tất cả\" active khi mở màn hình",
   "Lần đầu mở Màn hình [Danh sách lịch hẹn].",
   "1. Kiểm tra trạng thái tab khi màn hình vừa load xong.",
   "1. Tab \"Tất cả\" ở trạng thái active (text đỏ + underline đỏ); 5 tab còn lại inactive.")

tc("[Card] Kiểm tra hiển thị đầy đủ 5 field + icon điều hướng",
   "Danh sách có ≥1 card có đủ dữ liệu.",
   "1. Kiểm tra hiển thị 1 card lịch hẹn.",
   "1. Card hiển thị đầy đủ:\n- Tên thủ tục (bold, tối đa 2 dòng).\n- Badge trạng thái.\n- Icon Tòa nhà (xám) + \"Lĩnh vực: \" + Tên lĩnh vực (1 dòng).\n- Icon Người (xám) + \"Thời gian đặt: \" + DD/MM/YYYY HH:mm.\n- Icon Lịch (xám) + \"Ngày cán bộ hẹn nộp: \" + DD/MM/YYYY.\n- Icon mũi tên \">\" xám góc phải.")

tc("[Card] Kiểm tra Tên thủ tục dài → truncate 2 dòng + \"...\"",
   "Có lịch hẹn với Tên thủ tục > 2 dòng.",
   "1. Kiểm tra hiển thị trường Tên thủ tục trên card.",
   "1. Tên thủ tục hiển thị tối đa 2 dòng; phần vượt quá bị cắt và hiển thị \"...\" ở cuối dòng 2.",
   "BVA")

tc("[Card] Kiểm tra Lĩnh vực dài → truncate 1 dòng + \"...\"",
   "Có lịch hẹn với Lĩnh vực > 1 dòng.",
   "1. Kiểm tra hiển thị trường Lĩnh vực trên card.",
   "1. Lĩnh vực hiển thị tối đa 1 dòng; phần vượt quá bị cắt và hiển thị \"...\" ở cuối.",
   "BVA")

tc("[Card] Kiểm tra field null → hiển thị \"-\"",
   "Có lịch hẹn với Lĩnh vực = null và Ngày cán bộ hẹn nộp = null.",
   "1. Kiểm tra hiển thị các field null trên card.",
   "1.\n- Card hiển thị \"Lĩnh vực: -\".\n- Card hiển thị \"Ngày cán bộ hẹn nộp: -\".\n(AC3 — Null handling)")

tc("[Card] Kiểm tra format \"Thời gian đặt\" DD/MM/YYYY HH:mm (24h, GMT+7)",
   "Có lịch hẹn với \"Thời gian đặt\" = 08/05/2026 14:30 (GMT+7).",
   "1. Kiểm tra hiển thị field \"Thời gian đặt\" trên card.",
   "1. Hiển thị chính xác \"Thời gian đặt: 08/05/2026 14:30\" (CMR-12).")

tc("[Card] Kiểm tra format \"Ngày cán bộ hẹn nộp\" DD/MM/YYYY (không có giờ)",
   "Có lịch hẹn với Ngày cán bộ hẹn nộp = 10/05/2026.",
   "1. Kiểm tra hiển thị field \"Ngày cán bộ hẹn nộp\" trên card.",
   "1. Hiển thị \"Ngày cán bộ hẹn nộp: 10/05/2026\" — không kèm giờ (CMR-12).")

tc("[Card] Kiểm tra Badge trạng thái hiển thị đúng màu follow UI design",
   "Danh sách có card với các trạng thái khác nhau.",
   "1. Kiểm tra hiển thị Badge cho từng trạng thái: \"Chờ xác nhận\", \"Đã xác nhận\", \"Đã hủy\", \"Đã bỏ lượt\", \"Đã hoàn thành\".",
   "1. Badge hiển thị đúng màu sắc follow theo UI design cho từng trạng thái (AC6, CMR-05).",
   "Follow UI design")

tc("[Bottom Sheet Filter] Kiểm tra UI/UX khi mở Filter",
   "Đang ở Màn hình [Danh sách lịch hẹn].",
   "1. Nhấp icon [Filter].\n2. Kiểm tra hiển thị Bottom Sheet.",
   "2.\n- Bottom Sheet trượt lên từ dưới.\n- Hiển thị đầy đủ: Dropdown \"Lĩnh vực\" (default \"Tất cả lĩnh vực\"), Dropdown \"Dịch vụ công\" (default \"Tất cả dịch vụ công\"), nút [X] góc phải trên, nút [Nhập lại] outline đỏ, nút [Áp dụng] filled đỏ.\n(Tham khảo ảnh \"UC42-filter\" sheet WFDesign).")

# --- Check Function ---
check_type("Check Function")

tc("[Entry point Quick Access] Kiểm tra truy cập từ Quick Access \"Quản lý đặt lịch\" trên Trang chủ",
   "Đăng nhập thành công, đang ở Trang chủ.",
   "1. Nhấp Quick Access \"Quản lý đặt lịch\" trên Trang chủ.",
   "1. Mở Màn hình [Danh sách lịch hẹn] với Tab \"Tất cả\" active mặc định.")

tc("[Entry point Sidebar] Kiểm tra truy cập từ Sidebar \"Quản lý đặt lịch\"",
   "Đăng nhập thành công, Sidebar đang mở.",
   "1. Nhấp item \"Quản lý đặt lịch\" trên Sidebar.",
   "1. Đóng Sidebar.\n2. Mở Màn hình [Danh sách lịch hẹn] giống hệt entry từ Quick Access (tab \"Tất cả\" active).")

tc("[Nút Back ←] Kiểm tra tap Back trên Header → quay về màn trước",
   "Đang ở Màn hình [Danh sách lịch hẹn] (đi từ Trang chủ).",
   "1. Tap nút [Back ←] trên Header.",
   "1. Đóng màn [Danh sách lịch hẹn], quay về Trang chủ (CMR-06).")

tc("[Role Cá nhân] Kiểm tra tài khoản Cá nhân chỉ thấy lịch hẹn của bản thân",
   "Đăng nhập tài khoản Cá nhân.",
   "1. Mở Màn hình [Danh sách lịch hẹn].\n2. Kiểm tra danh sách lịch hẹn.",
   "2. Danh sách chỉ chứa lịch hẹn của tài khoản Cá nhân này; không thấy lịch hẹn của tài khoản/tổ chức khác.",
   "Role-based")

tc("[Role Tổ chức] Kiểm tra tài khoản Tổ chức chỉ thấy lịch hẹn của tổ chức",
   "Đăng nhập tài khoản Tổ chức.",
   "1. Mở Màn hình [Danh sách lịch hẹn].\n2. Kiểm tra danh sách lịch hẹn.",
   "2. Danh sách chỉ chứa lịch hẹn của tổ chức này; không thấy lịch hẹn của tổ chức/cá nhân khác.",
   "Role-based")

tc("[Guest] Kiểm tra chưa đăng nhập truy cập UC42-44 → redirect đăng nhập",
   "Chưa đăng nhập (hoặc đã logout).",
   "1. Cố gắng truy cập Màn hình [Danh sách lịch hẹn] (qua deep-link nếu có, hoặc app ẩn Quick Access/Sidebar item).",
   "1. App ẩn entry point hoặc redirect về màn [Đăng nhập]; không cho xem danh sách.",
   "Permission")

tc("[Tab \"Chờ xác nhận\"] Kiểm tra lọc danh sách theo trạng thái",
   "Đang ở Tab \"Tất cả\", có ≥1 lịch hẹn trạng thái \"Chờ xác nhận\".",
   "1. Tap Tab \"Chờ xác nhận\".",
   "1.\n- Tab \"Tất cả\" unselect (bỏ underline đỏ).\n- Tab \"Chờ xác nhận\" active (text đỏ + underline đỏ).\n- Danh sách chỉ hiển thị lịch hẹn trạng thái \"Chờ xác nhận\".",
   "EP")

tc("[Tab \"Đã xác nhận\"] Kiểm tra lọc danh sách theo trạng thái",
   "Đang ở Tab \"Tất cả\", có ≥1 lịch hẹn trạng thái \"Đã xác nhận\".",
   "1. Tap Tab \"Đã xác nhận\".",
   "1. Chỉ hiển thị lịch hẹn trạng thái \"Đã xác nhận\"; tab này active, các tab khác inactive.",
   "EP")

tc("[Tab \"Đã hủy\"] Kiểm tra lọc danh sách theo trạng thái",
   "Có ≥1 lịch hẹn trạng thái \"Đã hủy\".",
   "1. Tap Tab \"Đã hủy\".",
   "1. Chỉ hiển thị lịch hẹn trạng thái \"Đã hủy\".",
   "EP")

tc("[Tab \"Đã bỏ lượt\"] Kiểm tra lọc danh sách theo trạng thái",
   "Có ≥1 lịch hẹn trạng thái \"Đã bỏ lượt\".",
   "1. Tap Tab \"Đã bỏ lượt\".",
   "1. Chỉ hiển thị lịch hẹn trạng thái \"Đã bỏ lượt\".",
   "EP")

tc("[Tab \"Đã hoàn thành\"] Kiểm tra lọc danh sách theo trạng thái",
   "Có ≥1 lịch hẹn trạng thái \"Đã hoàn thành\".",
   "1. Tap Tab \"Đã hoàn thành\".",
   "1. Chỉ hiển thị lịch hẹn trạng thái \"Đã hoàn thành\".",
   "EP")

tc("[Tab bar] Kiểm tra single selection — chỉ 1 tab active tại 1 thời điểm",
   "Đang ở Màn hình [Danh sách lịch hẹn].",
   "1. Tap lần lượt các tab \"Chờ xác nhận\" → \"Đã xác nhận\" → \"Đã hủy\".\n2. Sau mỗi lần tap, kiểm tra trạng thái tab bar.",
   "1 & 2. Tại mọi thời điểm chỉ có đúng 1 tab active (text đỏ + underline); 5 tab còn lại inactive.")

tc("[Tab bar] Kiểm tra tab bar swipe ngang khi không hiển thị đủ 6 tab",
   "Thiết bị màn hình nhỏ không hiển thị đủ 6 tab cùng lúc.",
   "1. Vuốt ngang trên Tab bar.",
   "1. Tab bar cho phép scroll ngang để xem tab bị che khuất.")

tc("[Tab] Kiểm tra Empty state \"Không có dữ liệu.\" cho tab không có data",
   "Tab \"Đã hủy\" không có lịch hẹn nào.",
   "1. Tap Tab \"Đã hủy\".",
   "1. Vùng danh sách hiển thị empty state \"Không có dữ liệu.\" (CMR-14).")

tc("[Sort order] Kiểm tra danh sách sắp xếp theo \"Thời gian đặt\" giảm dần",
   "Tab \"Tất cả\" có nhiều lịch hẹn với \"Thời gian đặt\" khác nhau.",
   "1. Kiểm tra thứ tự các card trong danh sách.",
   "1. Các card sắp xếp theo \"Thời gian đặt\" giảm dần (bản ghi có thời gian đặt gần nhất ở trên cùng) — AC4.")

tc("[Ô Search] Kiểm tra happy path — nhập keyword hợp lệ → filter đúng sau 3s",
   "Đang ở Tab \"Tất cả\", có lịch hẹn chứa keyword \"đầu tư\".",
   "1. Nhập \"đầu tư\" vào ô [Search].\n2. Chờ 3 giây.\n3. Kiểm tra danh sách.",
   "3. Danh sách tự hiển thị kết quả chứa \"đầu tư\" (fuzzy match) mà không cần nhấn Enter (AC9, CMR-01).")

tc("[Ô Search] Kiểm tra debounce 3s — gõ liên tục không trigger filter giữa chừng",
   "Đang ở Tab \"Tất cả\".",
   "1. Gõ từng ký tự \"đ\", \"ầ\", \"u\", \"t\", \"ư\" liên tục trong < 3s.\n2. Kiểm tra danh sách trong khi đang gõ.\n3. Sau khi ngừng gõ 3s, kiểm tra lại.",
   "2. Trong quá trình gõ, danh sách KHÔNG trigger filter sau mỗi ký tự.\n3. Sau 3s kể từ ký tự cuối, danh sách mới filter 1 lần.")

tc("[Ô Search] Kiểm tra nhập đúng 1 ký tự (min boundary)",
   "Đang ở Tab \"Tất cả\".",
   "1. Nhập đúng 1 ký tự \"a\".\n2. Chờ 3s.",
   "2. Danh sách filter theo keyword 1 ký tự (fuzzy match).",
   "BVA min")

tc("[Ô Search] Kiểm tra nhập đúng 500 ký tự (max boundary)",
   "Đang ở Tab \"Tất cả\".",
   "1. Nhập chuỗi đúng 500 ký tự (lặp \"a\" x 500).\n2. Chờ 3s.\n3. Kiểm tra ô Search + danh sách.",
   "3. Ô [Search] chứa đúng 500 ký tự; danh sách trigger filter sau 3s (CMR-01).",
   "BVA max")

tc("[Ô Search] Kiểm tra chặn ký tự thứ 501 (max + 1)",
   "Ô [Search] đang có 500 ký tự.",
   "1. Cố gắng gõ thêm ký tự thứ 501.",
   "1. Ô [Search] KHÔNG nhận thêm ký tự; giữ nguyên 500 ký tự (hoặc auto truncate = 500).",
   "BVA max+1")

tc("[Ô Search] Kiểm tra paste chuỗi > 500 ký tự → cắt còn 500",
   "Clipboard chứa chuỗi 600 ký tự.",
   "1. Focus ô [Search].\n2. Paste.",
   "2. Ô [Search] chỉ nhận 500 ký tự đầu tiên; 100 ký tự còn lại bị bỏ.",
   "BVA paste")

tc("[Ô Search] Kiểm tra search theo \"Tên thủ tục\" — fuzzy match",
   "Có lịch hẹn với Tên thủ tục chứa \"cấp giấy\".",
   "1. Nhập \"cấp giấy\" vào ô [Search].\n2. Chờ 3s.",
   "2. Kết quả hiển thị tất cả lịch hẹn có Tên thủ tục chứa \"cấp giấy\".")

tc("[Ô Search] Kiểm tra search theo \"Mã thủ tục\" — fuzzy match",
   "Có lịch hẹn với Mã thủ tục chứa \"2.000412\".",
   "1. Nhập \"2.000412\" vào ô [Search].\n2. Chờ 3s.",
   "2. Kết quả hiển thị tất cả lịch hẹn có Mã thủ tục chứa \"2.000412\".")

tc("[Ô Search] Kiểm tra scope search áp dụng toàn bộ tab",
   "Đang ở Tab \"Chờ xác nhận\", có lịch hẹn ở các tab khác chứa keyword \"xyz-chung\".",
   "1. Nhập \"xyz-chung\" vào ô [Search].\n2. Chờ 3s.\n3. Kiểm tra kết quả + tab active.",
   "3.\n- Kết quả trả về bao gồm lịch hẹn từ TẤT CẢ trạng thái (không giới hạn \"Chờ xác nhận\").\n- Kết quả hiển thị trên Tab \"Tất cả\" (AC9).")

tc("[Ô Search] Kiểm tra keyword không có match → \"Không tìm thấy kết quả.\"",
   "Không có lịch hẹn nào match keyword \"xyz123notexist\".",
   "1. Nhập \"xyz123notexist\" vào ô [Search].\n2. Chờ 3s.",
   "2. Hiển thị empty state \"Không tìm thấy kết quả.\" (CMR-14).")

tc("[Ô Search] Kiểm tra xóa hết keyword → danh sách reset về mặc định",
   "Đã nhập keyword và có kết quả filter.",
   "1. Xóa toàn bộ ký tự trong ô [Search].\n2. Chờ 3s.",
   "2. Danh sách trở về trạng thái mặc định theo Tab đang chọn (hiển thị tất cả) — AC10.")

tc("[Ô Search] Kiểm tra nhập chỉ khoảng trắng → coi như rỗng",
   "Đang ở Tab \"Tất cả\".",
   "1. Nhập chỉ spaces \"   \".\n2. Chờ 3s.",
   "2. Hệ thống xử lý như rỗng; hiển thị danh sách mặc định (không filter vô nghĩa).",
   "EP invalid")

tc("[Ô Search] Kiểm tra nhập keyword tiếng Việt có dấu (UTF-8)",
   "Có lịch hẹn với Tên thủ tục chứa \"Đầu tư\".",
   "1. Nhập \"Đầu tư\" (có dấu).\n2. Chờ 3s.",
   "2. Search khớp đúng; các ký tự UTF-8 không gây lỗi.",
   "EP Unicode")

tc("[Ô Search] Kiểm tra nhập emoji/ký tự đặc biệt → không crash",
   "Đang ở Tab \"Tất cả\".",
   "1. Nhập \"😀\" hoặc \"<>&'\".\n2. Chờ 3s.",
   "2. App không crash; hiển thị \"Không tìm thấy kết quả.\" nếu không match.",
   "Security client-side")

tc("[Icon Filter] Kiểm tra tap mở Bottom Sheet",
   "Đang ở Màn hình [Danh sách lịch hẹn].",
   "1. Tap icon [Filter].",
   "1. Bottom Sheet Filter trượt lên từ dưới, hiển thị đầy đủ các control (CMR-02).")

tc("[Dropdown Lĩnh vực] Kiểm tra tap mở dropdown",
   "Bottom Sheet Filter đang mở.",
   "1. Tap Dropdown \"Lĩnh vực\".",
   "1. Mở danh sách option Lĩnh vực từ danh mục hệ thống; option \"Tất cả lĩnh vực\" ở trên cùng.")

tc("[Dropdown Lĩnh vực] Kiểm tra option đã chọn được highlight/bold khi mở lại",
   "Đã chọn Lĩnh vực X trước đó.",
   "1. Mở lại Dropdown \"Lĩnh vực\".\n2. Kiểm tra hiển thị các option.",
   "2. Option X được highlight/bold; các option khác hiển thị bình thường (CMR-03).")

tc("[Dropdown Lĩnh vực] Kiểm tra tìm kiếm inline trong dropdown",
   "Dropdown \"Lĩnh vực\" đang mở.",
   "1. Nhập vài ký tự vào ô tìm kiếm trong dropdown.",
   "1. Danh sách option tự lọc theo fuzzy match.")

tc("[Dropdown Lĩnh vực] Kiểm tra option text dài → truncate \"...\"",
   "Có option với tên vượt giới hạn ký tự hiển thị.",
   "1. Mở Dropdown \"Lĩnh vực\".\n2. Kiểm tra hiển thị option dài.",
   "2. Option tự cắt ngắn và thêm \"...\" ở cuối (CMR-03).",
   "BVA")

tc("[Filter] Kiểm tra chỉ chọn Lĩnh vực (Valid) + DVC mặc định",
   "Bottom Sheet Filter đang mở.",
   "1. Chọn Lĩnh vực \"Đầu tư\".\n2. Giữ Dịch vụ công = \"Tất cả dịch vụ công\".\n3. Tap [Áp dụng].",
   "3. Bottom Sheet đóng; danh sách filter theo Lĩnh vực \"Đầu tư\" (mọi dịch vụ công) — AC7.",
   "Decision Table")

tc("[Filter] Kiểm tra chỉ chọn Dịch vụ công (Valid) + Lĩnh vực mặc định",
   "Bottom Sheet Filter đang mở.",
   "1. Giữ Lĩnh vực = \"Tất cả lĩnh vực\".\n2. Chọn Dịch vụ công \"Dịch vụ công trực tuyến mức độ 4\".\n3. Tap [Áp dụng].",
   "3. Danh sách filter theo DVC \"mức độ 4\" (mọi lĩnh vực).",
   "Decision Table")

tc("[Filter] Kiểm tra chọn cả Lĩnh vực + Dịch vụ công",
   "Bottom Sheet Filter đang mở.",
   "1. Chọn Lĩnh vực \"Đầu tư\".\n2. Chọn Dịch vụ công \"Mức độ 4\".\n3. Tap [Áp dụng].",
   "3. Danh sách filter đồng thời theo CẢ 2 tiêu chí (AND logic) — AC7.",
   "Decision Table")

tc("[Filter] Kiểm tra combo không có data matching",
   "Bottom Sheet Filter đang mở; không có lịch hẹn nào match combo sau đây.",
   "1. Chọn Lĩnh vực X.\n2. Chọn Dịch vụ công Y (combo không có data).\n3. Tap [Áp dụng].",
   "3. Bottom Sheet đóng; hiển thị \"Không tìm thấy kết quả.\" (CMR-14).",
   "Decision Table")

tc("[Filter] Kiểm tra filter độc lập — chọn Lĩnh vực KHÔNG ảnh hưởng danh sách Dịch vụ công",
   "Bottom Sheet Filter đang mở.",
   "1. Mở Dropdown \"Dịch vụ công\", ghi nhận số lượng option.\n2. Đóng dropdown, chọn 1 Lĩnh vực X.\n3. Mở lại Dropdown \"Dịch vụ công\".",
   "3. Số lượng option Dịch vụ công KHÔNG bị cắt/thu hẹp theo Lĩnh vực X; vẫn đầy đủ như bước 1 (AC7).",
   "No cascade")

tc("[Nút Nhập lại] Kiểm tra reset filter, không đóng Bottom Sheet, không gọi API",
   "Đã chọn Lĩnh vực X + Dịch vụ công Y, Bottom Sheet đang mở.",
   "1. Tap nút [Nhập lại].",
   "1.\n- Dropdown Lĩnh vực reset về \"Tất cả lĩnh vực\".\n- Dropdown Dịch vụ công reset về \"Tất cả dịch vụ công\".\n- Bottom Sheet VẪN MỞ.\n- Danh sách phía sau KHÔNG thay đổi (không gọi API).")

tc("[Nút Áp dụng] Kiểm tra đóng Bottom Sheet, gọi API, cập nhật danh sách",
   "Đã chọn filter, Bottom Sheet đang mở.",
   "1. Tap nút [Áp dụng].",
   "1.\n- Bottom Sheet đóng.\n- App gọi API với tham số filter đã chọn.\n- Danh sách cập nhật theo filter.")

tc("[Active Filter Indicator] Kiểm tra indicator xanh lá khi filter ≠ mặc định",
   "Vừa Áp dụng filter có tiêu chí ≠ \"Tất cả\".",
   "1. Kiểm tra hiển thị icon [Filter].",
   "1. Icon [Filter] có chấm indicator màu xanh lá cây ở góc phải trên (AC8, CMR-02).")

tc("[Active Filter Indicator] Kiểm tra indicator ẩn khi reset về mặc định",
   "Đang có indicator xanh lá (filter active).",
   "1. Mở Filter.\n2. Tap [Nhập lại].\n3. Tap [Áp dụng].\n4. Kiểm tra icon [Filter].",
   "4. Indicator xanh lá ẩn đi; icon [Filter] về trạng thái mặc định.")

tc("[Nút X Bottom Sheet] Kiểm tra đóng Bottom Sheet, không áp dụng",
   "Bottom Sheet Filter đang mở, đã chọn tiêu chí mới.",
   "1. Tap nút [X] góc phải trên.",
   "1. Bottom Sheet đóng; danh sách phía sau giữ nguyên theo filter cũ; tiêu chí vừa chọn bị hủy (không lưu).")

tc("[Bottom Sheet] Kiểm tra tap vùng ngoài Bottom Sheet → đóng",
   "Bottom Sheet Filter đang mở.",
   "1. Tap vào vùng mờ ngoài Bottom Sheet.",
   "1. Bottom Sheet đóng, không thay đổi kết quả hiện tại.")

tc("[Card] Kiểm tra tap card → navigate sang Chi tiết lịch hẹn",
   "Đang ở Danh sách, có ≥1 card.",
   "1. Tap vào bất kỳ vùng nào của 1 card (hoặc icon \">\").",
   "1. Navigate sang Màn hình [Chi tiết lịch hẹn] với đúng mã lịch hẹn của card đó.")

tc("[Card] Kiểm tra Badge read-only — không có action riêng khi tap",
   "Đang ở Danh sách.",
   "1. Tap chính xác lên vùng Badge trạng thái của 1 card.",
   "1. Badge không có phản hồi riêng (không mở menu, không toast); app vẫn navigate sang [Chi tiết] vì tap toàn card.")

tc("[Card] Kiểm tra debounce double-tap — chỉ navigate 1 lần",
   "Đang ở Danh sách.",
   "1. Double-tap nhanh liên tục vào 1 card.",
   "1. App chỉ navigate 1 lần sang [Chi tiết]; không mở 2 instance (CMR-18).")

tc("[Lazy load] Kiểm tra first page load — 20 records đầu tiên",
   "Tài khoản có > 20 lịch hẹn, tab \"Tất cả\" active.",
   "1. Mở Màn hình [Danh sách lịch hẹn].\n2. Kiểm tra số lượng card hiển thị.",
   "2. Danh sách hiển thị tối đa 20 card (page 1) — AC13, CMR-04.")

tc("[Lazy load] Kiểm tra tổng 20 records → không trigger lazy load",
   "Tài khoản có đúng 20 lịch hẹn.",
   "1. Mở Danh sách.\n2. Cuộn xuống cuối danh sách.",
   "2.\n- Hiển thị đủ 20 card.\n- Không trigger thêm page; không hiển thị spinner lazy ở cuối.",
   "BVA size=20")

tc("[Lazy load] Kiểm tra tổng 21 records → cuộn cuối trigger page 2 (1 record)",
   "Tài khoản có đúng 21 lịch hẹn.",
   "1. Mở Danh sách (load page 1 = 20).\n2. Cuộn đến cuối.",
   "2. Auto trigger load page 2 thêm 1 record; tổng hiển thị 21.",
   "BVA size=21")

tc("[Lazy load] Kiểm tra tổng 40 records → 2 page đầy",
   "Tài khoản có đúng 40 lịch hẹn.",
   "1. Mở Danh sách.\n2. Cuộn đến cuối page 1.\n3. Cuộn đến cuối page 2.",
   "2. Trigger load page 2 (20 thêm).\n3. Không trigger thêm page; tổng 40.",
   "BVA size=40")

tc("[Lazy load] Kiểm tra spinner cục bộ khi load page tiếp theo",
   "Đã load page 1, tổng > 20 records.",
   "1. Cuộn đến cuối page 1.\n2. Kiểm tra hiển thị trong khi đang load page 2.",
   "2. Hiển thị spinner cục bộ ở cuối danh sách; spinner ẩn sau khi load xong (CMR-07).")

tc("[Lazy load] Kiểm tra retry tự động 3 lần khi page N fail",
   "Simulate API lazy load page 2 trả lỗi.",
   "1. Cuộn cuối trigger lazy load page 2.\n2. Kiểm tra hệ thống retry.",
   "2. App tự retry 3 lần (mỗi lần ~2s). Nếu 1 trong 3 lần thành công → append records; nếu tất cả fail → dừng.")

tc("[Lazy load] Kiểm tra sau 3 lần retry vẫn fail → hiển thị lỗi cục bộ",
   "Simulate API lazy load luôn fail.",
   "1. Cuộn cuối trigger lazy load.\n2. Chờ đủ 3 lần retry.",
   "2. Sau 3 lần fail: dừng auto retry; hiển thị message lỗi cục bộ ở cuối danh sách (AC14).")

tc("[Lazy load] Kiểm tra pull-to-refresh để tải lại sau lazy fail",
   "Đã hiển thị lỗi lazy load cục bộ ở cuối danh sách.",
   "1. Kéo xuống pull-to-refresh.",
   "1. App gọi lại API từ page 1; reset toàn bộ danh sách; ẩn message lỗi cục bộ.")

tc("[Pull to refresh] Kiểm tra trên Danh sách → refresh từ page 1",
   "Đang ở Danh sách có data.",
   "1. Kéo xuống từ đầu danh sách.\n2. Quan sát hiển thị.",
   "2.\n- Spinner hiển thị ở đầu danh sách.\n- App gọi API page 1.\n- Cập nhật danh sách.\n- Ẩn spinner (AC11, CMR-13).")

tc("[Pull to refresh] Kiểm tra giữ nguyên tab + search + filter hiện tại",
   "Đang Tab \"Chờ xác nhận\" + search keyword \"abc\" + filter active.",
   "1. Pull-to-refresh.",
   "1. Refresh nhưng vẫn giữ tab \"Chờ xác nhận\", keyword \"abc\", filter active.")

tc("[Pull to refresh] Kiểm tra khi đang ở page N → reset về page 1",
   "Đã load đến page 3 (>40 records hiển thị), scroll đang ở giữa.",
   "1. Pull-to-refresh.",
   "1. Danh sách load lại từ page 1 (tối đa 20 records); scroll về đầu; lazy state reset.")

tc("[Loading state] Kiểm tra first-load → full-screen overlay",
   "Lần đầu vào Màn hình [Danh sách lịch hẹn].",
   "1. Tap Quick Access.\n2. Kiểm tra hiển thị trong khi chờ API.",
   "2. Full-screen loading overlay hiển thị trên toàn màn hình; sau khi API trả → ẩn overlay, hiển thị danh sách (AC17).")

tc("[Loading state] Kiểm tra đổi tab → chỉ spinner cục bộ (không full-screen)",
   "Đã first-load xong, đang Tab \"Tất cả\".",
   "1. Tap Tab \"Chờ xác nhận\".\n2. Kiểm tra hiển thị trong khi chờ API.",
   "2. Chỉ hiển thị spinner cục bộ trong vùng danh sách; KHÔNG full-screen overlay.")

tc("[State persistence] Kiểm tra giữ nguyên tab khi quay lại từ Chi tiết",
   "Đang Tab \"Chờ xác nhận\" trên Danh sách.",
   "1. Tap 1 card.\n2. Vào Chi tiết.\n3. Tap Back (←).",
   "3. Quay về Danh sách với Tab \"Chờ xác nhận\" vẫn active (không reset về \"Tất cả\") — AC2.")

tc("[State persistence] Kiểm tra giữ scroll position khi quay lại từ Chi tiết",
   "Cuộn đến card thứ 15 trên Danh sách.",
   "1. Tap card 15.\n2. Vào Chi tiết.\n3. Tap Back (←).",
   "3. Quay về Danh sách với scroll vẫn ở vị trí card 15 (AC2).")

tc("[State persistence] Kiểm tra giữ search keyword khi quay lại từ Chi tiết",
   "Đã search keyword \"đầu tư\" trên Danh sách.",
   "1. Tap 1 card trong kết quả.\n2. Vào Chi tiết.\n3. Tap Back (←).",
   "3. Ô [Search] vẫn giữ \"đầu tư\"; danh sách vẫn hiển thị kết quả search (AC18, CMR-01).")

tc("[State persistence] Kiểm tra giữ filter khi quay lại từ Chi tiết",
   "Đã áp dụng filter Lĩnh vực = X.",
   "1. Tap 1 card trong kết quả.\n2. Vào Chi tiết.\n3. Tap Back (←).",
   "3. Indicator xanh lá trên icon [Filter] vẫn hiển thị; danh sách vẫn filter theo Lĩnh vực X (AC18).")

tc("[State persistence] Kiểm tra reset search + filter khi chuyển tab khác",
   "Tab \"Tất cả\" có search \"đầu tư\" + filter Lĩnh vực X active.",
   "1. Tap Tab \"Đã xác nhận\".",
   "1.\n- Ô [Search] rỗng (reset).\n- Filter indicator ẩn (reset về mặc định) — AC19.")

tc("[Physical Back Android] Kiểm tra Physical Back đóng Bottom Sheet khi đang mở",
   "Android, Bottom Sheet Filter đang mở.",
   "1. Nhấn Physical Back button trên thiết bị Android.",
   "1. Bottom Sheet đóng; không thoát màn [Danh sách]; danh sách không đổi (assumption Q6).",
   "Platform")

tc("[Error] Kiểm tra mất mạng (Danh sách) → message + nút Thử lại",
   "Đã đăng nhập, tắt mạng trước khi mở màn.",
   "1. Mở Màn hình [Danh sách lịch hẹn].",
   "1. Hiển thị message \"Không thể kết nối. Vui lòng kiểm tra mạng và thử lại.\" kèm nút \"Thử lại\" (AC15, CMR-07).")

tc("[Error] Kiểm tra lỗi 500 → message không có Thử lại",
   "Simulate API trả HTTP 500.",
   "1. Mở Danh sách.",
   "1. Hiển thị \"Hệ thống đang bận. Vui lòng thử lại sau.\" — KHÔNG có nút Thử lại (AC15, CMR-07).")

tc("[Error] Kiểm tra timeout > 10s → message + Thử lại",
   "Simulate API mất > 10s chưa phản hồi.",
   "1. Mở Danh sách.\n2. Chờ 10s.",
   "2. Hiển thị \"Yêu cầu đã hết thời gian chờ. Vui lòng thử lại.\" kèm nút \"Thử lại\" (CMR-16).")

tc("[Error] Kiểm tra API response 9.9s (dưới 10s) → load thành công",
   "Simulate API response ~9.9s.",
   "1. Mở Danh sách.\n2. Chờ đến khi response trả về.",
   "2. Danh sách hiển thị thành công; KHÔNG trigger timeout.",
   "BVA 9.9s")

tc("[Error] Kiểm tra API response 10.1s (trên 10s) → trigger timeout",
   "Simulate API response ~10.1s.",
   "1. Mở Danh sách.\n2. Chờ qua 10s.",
   "2. Trigger timeout; hiển thị \"Yêu cầu đã hết thời gian chờ. Vui lòng thử lại.\" kèm nút \"Thử lại\".",
   "BVA 10.1s")

tc("[Error 401] Kiểm tra 401 + refresh token hợp lệ → silent retry thành công",
   "API trả 401 nhưng refresh token còn hiệu lực (< 15 ngày).",
   "1. Mở Danh sách.\n2. Kiểm tra hành vi phía client.",
   "2. App tự động refresh token và retry API gốc; user không bị gián đoạn; danh sách hiển thị thành công.")

tc("[Error 401] Kiểm tra 401 + refresh token hết hạn (>15 ngày) → toast + redirect login",
   "API trả 401, refresh token cũng đã hết hạn.",
   "1. Mở Danh sách.",
   "1. Hiển thị toast \"Phiên đăng nhập hết hạn.\"; redirect về màn [Đăng nhập] (AC16, CMR-07).")

tc("[Error Retry] Kiểm tra tap \"Thử lại\" sau lỗi mạng + mạng phục hồi → load thành công",
   "Đang hiển thị lỗi mạng với nút \"Thử lại\".",
   "1. Bật mạng lại.\n2. Tap \"Thử lại\".",
   "2. App gọi lại API thành công; ẩn lỗi; hiển thị danh sách.")

tc("[Offline] Kiểm tra offline khi đang xem danh sách → pull-to-refresh báo lỗi",
   "Đang xem danh sách có data.",
   "1. Tắt mạng.\n2. Pull-to-refresh.",
   "2. Hiển thị lỗi mạng (không cache theo assumption Q30).",
   "Assumption")

tc("[i18n] Kiểm tra đổi ngôn ngữ sang English → text cứng đổi, text động giữ nguyên",
   "Ngôn ngữ hiện tại VI. Có ≥1 lịch hẹn.",
   "1. Đổi ngôn ngữ app sang English.\n2. Mở Màn hình [Danh sách lịch hẹn].\n3. Kiểm tra text.",
   "3.\n- Text cứng đổi sang EN: Header, tên tab, label, placeholder, nút, message.\n- Text động giữ nguyên: Tên thủ tục, Tên lĩnh vực, Tên dịch vụ công (từ API) — AC20, CMR-17.")

tc("[i18n] Kiểm tra đổi ngôn ngữ sang Tiếng Trung → text cứng đổi",
   "Ngôn ngữ hiện tại VI.",
   "1. Đổi ngôn ngữ app sang 中文.\n2. Mở Danh sách.\n3. Kiểm tra text.",
   "3. Text cứng (Header, tab, label, button, message) đổi sang Tiếng Trung; text động giữ nguyên.")

tc("[i18n] Kiểm tra đổi ngôn ngữ sang Tiếng Nhật → text cứng đổi",
   "Ngôn ngữ hiện tại VI.",
   "1. Đổi ngôn ngữ app sang 日本語.\n2. Mở Danh sách.\n3. Kiểm tra text.",
   "3. Text cứng đổi sang Tiếng Nhật; text động giữ nguyên.")

tc("[i18n] Kiểm tra đổi ngôn ngữ sang Tiếng Hàn → text cứng đổi",
   "Ngôn ngữ hiện tại VI.",
   "1. Đổi ngôn ngữ app sang 한국어.\n2. Mở Danh sách.",
   "Text cứng đổi sang Tiếng Hàn; text động giữ nguyên.")

tc("[i18n] Kiểm tra format ngày giờ giữ nguyên DD/MM/YYYY HH:mm GMT+7 trên mọi locale",
   "Đã đổi ngôn ngữ lần lượt VI → EN → JA → KO.",
   "1. Tại mỗi ngôn ngữ, quan sát format \"Thời gian đặt\" trên card.",
   "Format vẫn là DD/MM/YYYY HH:mm (24h GMT+7) — KHÔNG đổi sang MM/DD/YYYY hay 12h AM/PM (CMR-12).")

tc("[A11y] Kiểm tra Screen reader đọc được Header, Tab, Card, Button",
   "Bật VoiceOver (iOS) hoặc TalkBack (Android).",
   "1. Điều hướng qua Header, Tab bar, Card, icon Filter, nút Back.",
   "1. Mỗi thành phần được đọc rõ ràng theo nội dung (AC21).")

tc("[Security] Kiểm tra search box không bị XSS/SQL injection qua UI",
   "Đang ở Danh sách.",
   "1. Nhập <script>alert(1)</script> vào ô [Search].\n2. Chờ 3s.",
   "2. App không thực thi script; hiển thị \"Không tìm thấy kết quả.\" hoặc coi như keyword bình thường; không crash.",
   "Security client-side")

tc("[UX Loading] Kiểm tra không có overlay khi chỉ switch tab",
   "First-load đã hoàn tất.",
   "1. Tap Tab bất kỳ khác.",
   "1. Không có full-screen overlay; chỉ có spinner cục bộ trong vùng danh sách.")

# --- Check common ---
check_type("Check common")

sub_label("Kiểm tra các trường hợp phổ biến UI/UX")

tc("Kiểm tra hiển thị dữ liệu tối đa (maxlength)",
   "",
   "1. Kiểm tra hiển thị dữ liệu tối đa (maxlength)",
   "1. Hiển thị đúng độ dài tối đa")

tc("Kiểm tra khôi phục/phóng to/thu nhỏ ứng dụng",
   "",
   "1. Thực hiện khôi phục, phóng to, thu nhỏ ứng dụng",
   "1. Không xảy ra lỗi bất thường")

tc("Kiểm tra tính nhất quán của các thông báo",
   "",
   "1. Kiểm tra tính nhất quán của các thông báo",
   "1. Xác nhận thông báo lỗi:\n- Thông báo lỗi inline: hiển thị dưới ô nhập liệu bị lỗi, màu đỏ\n- Thông báo lỗi dạng toast: hiển thị ở giữa hoặc phía dưới màn hình")

tc("Kiểm tra hiển thị khi thiết bị ở chế độ dọc (thiết bị xoay theo chiều đứng)",
   "",
   "1. Kiểm tra hiển thị khi thiết bị ở chế độ dọc",
   "1. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị ở chế độ ngang (thiết bị xoay theo chiều ngang)",
   "",
   "1. Kiểm tra hiển thị khi thiết bị ở chế độ ngang",
   "1. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi chuyển đổi giữa chế độ dọc và ngang",
   "",
   "1. Kiểm tra hiển thị khi chuyển từ chế độ dọc sang ngang\n2. Kiểm tra hiển thị khi chuyển từ chế độ ngang sang dọc",
   "1 & 2. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ lớn nhất",
   "",
   "1. Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ lớn nhất",
   "1. Giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ nhỏ nhất",
   "",
   "1. Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ nhỏ nhất",
   "1. Giao diện không bị vỡ")

sub_label("Kiểm tra tương tác cơ bản với thiết bị")

tc("Xác nhận hiển thị của ứng dụng khi người dùng chạm vào nút [Quay lại] trên thiết bị Android",
   "",
   "1. Mở ứng dụng\n2. Người dùng chạm vào nút [Quay lại] trên thiết bị Android\n3. Xác nhận hiển thị",
   "3. Quay lại màn hình trước đó")

tc("Xác nhận hiển thị của ứng dụng khi người dùng vuốt từ trái sang phải trên thiết bị iOS",
   "",
   "1. Mở ứng dụng\n2. Người dùng vuốt từ trái sang phải trên thiết bị iOS\n3. Xác nhận hiển thị",
   "3. Quay lại màn hình trước đó")

tc("Xác nhận hiển thị của ứng dụng khi người dùng tắt và mở lại ứng dụng",
   "",
   "1. Mở ứng dụng\n2. Người dùng tắt ứng dụng\n3. Mở lại ứng dụng\n4. Xác nhận hiển thị",
   "4. Ứng dụng mở lại từ trạng thái ban đầu")

tc("Kiểm tra chế độ đa nhiệm (multitasking)",
   "",
   "1. Mở ứng dụng\n2. Trở về màn hình chính (không tắt ứng dụng)\n3. Mở lại ứng dụng\n4. Xác nhận hiển thị",
   "4. Ứng dụng giữ nguyên ở trạng thái hiện tại")

tc("Xác nhận hiển thị của ứng dụng khi người dùng khóa và mở khóa màn hình thiết bị",
   "",
   "1. Mở ứng dụng\n2. Khóa thiết bị\n3. Mở khóa thiết bị\n4. Xác nhận hiển thị",
   "4. Giữ nguyên trạng thái hiện tại của ứng dụng")

tc("Kiểm tra hành động kéo xuống để làm mới (pull-to-refresh)",
   "Tiền điều kiện: Màn hình hỗ trợ tính năng kéo để làm mới",
   "1. Người dùng ở màn hình hiện tại\n2. Kéo xuống để làm mới",
   "2. Hiển thị dữ liệu mới nhất của màn hình")

tc("Kiểm tra hành động cuộn xuống để tải thêm (scroll-down-to-load-more)",
   "Tiền điều kiện: Màn hình hỗ trợ tính năng cuộn xuống để tải thêm",
   "1. Người dùng cuộn xuống cuối danh sách",
   "2. Hiển thị thêm dữ liệu mới")

tc("Kiểm tra phản hồi của ứng dụng khi thiết bị nhận được thông báo từ ứng dụng khác",
   "Tiền điều kiện: Ứng dụng khác được phép gửi thông báo",
   "1. Người dùng mở ứng dụng\n2. Ứng dụng khác gửi thông báo\n3. Xác nhận hiển thị",
   "3. Không có lỗi nào xảy ra")

# ============================================================
# SECTION 2: Màn hình Chi tiết Lịch hẹn
# ============================================================
section("2. Màn hình Chi tiết Lịch hẹn")

# --- Check UI/UX ---
check_type("Check UI/UX")

tc("[Màn hình Chi tiết lịch hẹn] Kiểm tra UI/UX khi có đủ dữ liệu",
   "Đăng nhập, có ≥1 lịch hẹn có đầy đủ 5 section dữ liệu.",
   "1. Tap 1 card bất kỳ trên Danh sách.\n2. Kiểm tra hiển thị Màn hình [Chi tiết lịch hẹn].",
   "2.\n- Hiển thị đầy đủ các item, màu sắc, font chữ, layout giống Design\n(Tham khảo ảnh \"UC42-44. Chi tiết lịch hẹn\" sheet WFDesign)\n- Header đỏ: nút Back (←) trái, tiêu đề \"Chi tiết lịch hẹn\" trắng căn giữa.\n- 5 Section theo thứ tự: (1) Thông tin thủ tục, (2) Thông tin người nộp, (3) Thông tin lịch hẹn, (4) Trạng thái, (5) Ghi chú.\n- Toàn bộ read-only, không có CTA.")

tc("[Màn hình Chi tiết lịch hẹn] Kiểm tra Loading state toàn màn hình khi first-load",
   "Tap 1 card đang trong quá trình chờ API chi tiết.",
   "1. Tap card.\n2. Kiểm tra hiển thị trong khi chờ API trả dữ liệu chi tiết.",
   "2. Full-screen loading overlay hiển thị trên toàn vùng nội dung (AC17, CMR-07); sau khi API trả → ẩn overlay.")

tc("[Section 1 - Thông tin thủ tục] Kiểm tra hiển thị đủ 6 field",
   "Mở Chi tiết của lịch hẹn có đầy đủ dữ liệu.",
   "1. Kiểm tra hiển thị Section 1.",
   "1. Section 1 hiển thị đúng 6 field:\n- Mã thủ tục\n- Tên thủ tục\n- Lĩnh vực\n- Dịch vụ công\n- Cơ quan thực hiện\n- Đơn vị tiếp nhận\nTất cả read-only.")

tc("[Section 2 - Thông tin người nộp] Kiểm tra hiển thị đủ 2 field",
   "Mở Chi tiết.",
   "1. Kiểm tra hiển thị Section 2.",
   "1. Section 2 hiển thị đúng 2 field: \"Mã định danh\" + \"Tên người nộp\"; đều read-only.")

tc("[Section 3 - Thông tin lịch hẹn] Kiểm tra hiển thị đủ 4 field",
   "Mở Chi tiết.",
   "1. Kiểm tra hiển thị Section 3.",
   "1. Section 3 hiển thị đúng 4 field: Ngày hẹn nộp, Khung giờ hẹn nộp, Thời gian đặt, Ngày cán bộ hẹn nộp; đều read-only.")

tc("[Section 4 - Trạng thái] Kiểm tra hiển thị Badge trạng thái",
   "Mở Chi tiết.",
   "1. Kiểm tra hiển thị Section 4.",
   "1. Section 4 hiển thị duy nhất 1 field \"Trạng thái\" dạng Badge; màu sắc follow UI design (CMR-05, AC6).",
   "Follow UI design")

tc("[Section 5 - Ghi chú] Kiểm tra hiển thị \"Ghi chú\"",
   "Mở Chi tiết của lịch hẹn có nội dung ghi chú.",
   "1. Kiểm tra hiển thị Section 5.",
   "1. Section 5 hiển thị nội dung ghi chú dạng text; wrap text nếu dài, KHÔNG truncate.")

tc("[Null handling] Kiểm tra các field null hiển thị \"-\"",
   "Mở Chi tiết của lịch hẹn có các field null (Mã thủ tục, Ngày hẹn nộp, Khung giờ hẹn nộp, Ngày cán bộ hẹn nộp, Ghi chú).",
   "1. Kiểm tra các field null trong Chi tiết.",
   "1. Các field null hiển thị \"-\" thay thế.")

tc("[Wrap text] Kiểm tra field nội dung dài → wrap, KHÔNG truncate",
   "Có lịch hẹn với Tên thủ tục / Lĩnh vực / Cơ quan thực hiện / Ghi chú rất dài.",
   "1. Kiểm tra hiển thị các field dài trong Chi tiết.",
   "1. Text wrap xuống nhiều dòng; KHÔNG có dấu \"...\" (Section 2.2 — Quy tắc hiển thị chung).")

tc("[Format ngày] Kiểm tra \"Ngày hẹn nộp\" hiển thị DD/MM/YYYY",
   "Lịch hẹn có Ngày hẹn nộp = 10/05/2026.",
   "1. Kiểm tra hiển thị field \"Ngày hẹn nộp\".",
   "1. Hiển thị đúng \"10/05/2026\" format DD/MM/YYYY (CMR-12).")

tc("[Format giờ] Kiểm tra \"Khung giờ hẹn nộp\" format \"HH:mm - HH:mm\"",
   "Lịch hẹn có Khung giờ hẹn nộp = \"08:00 - 09:00\".",
   "1. Kiểm tra hiển thị field \"Khung giờ hẹn nộp\".",
   "1. Hiển thị \"08:00 - 09:00\" (24h, GMT+7, CMR-12); KHÔNG ở dạng \"8h-9h\" hay \"08:00 AM - 09:00 AM\".")

tc("[Format ngày giờ] Kiểm tra \"Thời gian đặt\" hiển thị DD/MM/YYYY HH:mm",
   "Lịch hẹn có Thời gian đặt = \"08/05/2026 14:30\".",
   "1. Kiểm tra hiển thị field \"Thời gian đặt\".",
   "1. Hiển thị \"08/05/2026 14:30\" (24h, GMT+7, CMR-12).")

# --- Check Function ---
check_type("Check Function")

tc("[Nút Back ←] Kiểm tra tap Back → quay về Danh sách",
   "Đang ở Chi tiết (navigate từ Danh sách).",
   "1. Tap nút [Back ←] trên Header Chi tiết.",
   "1. Đóng Chi tiết; quay về Màn hình [Danh sách lịch hẹn] (CMR-06).")

tc("[Physical Back Android] Kiểm tra Physical Back từ Chi tiết → Danh sách",
   "Android, đang ở Chi tiết.",
   "1. Nhấn Physical Back button.",
   "1. Đóng Chi tiết; quay về Danh sách (giả định đồng nhất icon ← — Q6).",
   "Platform")

tc("[Data - Cá nhân] Kiểm tra \"Mã định danh\" hiển thị CCCD/CMND",
   "Đăng nhập tài khoản Cá nhân; mở Chi tiết 1 lịch hẹn của Cá nhân này.",
   "1. Kiểm tra Section 2 — field \"Mã định danh\".",
   "1. Hiển thị số CCCD/CMND đúng với dữ liệu API trả về (Section 2.2 Section 2).",
   "Role")

tc("[Data - Tổ chức] Kiểm tra \"Mã định danh\" hiển thị Mã doanh nghiệp",
   "Đăng nhập tài khoản Tổ chức; mở Chi tiết 1 lịch hẹn của Tổ chức này.",
   "1. Kiểm tra Section 2 — field \"Mã định danh\".",
   "1. Hiển thị Mã doanh nghiệp đúng với dữ liệu API trả về.",
   "Role")

tc("[Section 1] Kiểm tra dữ liệu Section 1 khớp 100% với API",
   "Mở Chi tiết 1 lịch hẹn X.",
   "1. So sánh 6 field Section 1 với response API chi tiết cho lịch hẹn X.",
   "1. Mã thủ tục / Tên thủ tục / Lĩnh vực / Dịch vụ công / Cơ quan thực hiện / Đơn vị tiếp nhận — khớp 100% với API (AC5).")

tc("[Section 2] Kiểm tra dữ liệu Section 2 khớp 100% với API",
   "Mở Chi tiết 1 lịch hẹn X.",
   "1. So sánh 2 field Section 2 với API.",
   "1. Mã định danh + Tên người nộp khớp 100% với API.")

tc("[Section 3] Kiểm tra dữ liệu Section 3 khớp 100% với API",
   "Mở Chi tiết 1 lịch hẹn X.",
   "1. So sánh 4 field Section 3 với API.",
   "1. Ngày hẹn nộp + Khung giờ + Thời gian đặt + Ngày cán bộ hẹn nộp khớp 100% với API; đúng format CMR-12.")

tc("[Consistency] Kiểm tra \"Thời gian đặt\" trên Chi tiết khớp 100% với card",
   "Đang ở Danh sách, ghi nhận \"Thời gian đặt\" trên card X.",
   "1. Tap card X.\n2. Kiểm tra field \"Thời gian đặt\" trên Chi tiết.",
   "2. Giá trị \"Thời gian đặt\" trên Chi tiết KHỚP 100% với trên card (cùng format DD/MM/YYYY HH:mm) — AC5.",
   "Integration")

tc("[Badge] Kiểm tra Badge read-only trong Chi tiết (không tap được)",
   "Đang ở Chi tiết.",
   "1. Tap vào vùng Badge Section 4.",
   "1. Không có phản hồi (không mở menu, không toast); Badge read-only thuần (AC6).")

tc("[Read-only scope] Kiểm tra toàn bộ Chi tiết read-only, không có CTA",
   "Đang ở Chi tiết.",
   "1. Quan sát toàn màn.",
   "1. Không có nút \"Hủy lịch\" / \"Chỉnh sửa\" / \"Tạo mới\" / bất kỳ CTA nào; chỉ có nút Back và gesture pull-to-refresh (Section 1 Out of scope + Section 2.2).")

tc("[Copy-to-clipboard] Kiểm tra field read-only không hỗ trợ copy",
   "Đang ở Chi tiết.",
   "1. Long-press/tap và giữ vào field bất kỳ (ví dụ Mã thủ tục).",
   "1. Không hiện menu copy-to-clipboard (giả định Q20).",
   "Assumption")

tc("[Pull to refresh] Kiểm tra pull-to-refresh trên Chi tiết → refresh dữ liệu",
   "Đang ở Chi tiết.",
   "1. Kéo xuống từ đầu màn.\n2. Quan sát hiển thị.",
   "2.\n- Spinner hiển thị ở đầu màn.\n- App gọi API chi tiết.\n- Cập nhật 5 section.\n- Ẩn spinner (AC12, CMR-13).")

tc("[Error] Kiểm tra mất mạng (Chi tiết) → message + Thử lại",
   "Tắt mạng.",
   "1. Tap 1 card.\n2. Quan sát hiển thị trên Chi tiết.",
   "2. \"Không thể kết nối. Vui lòng kiểm tra mạng và thử lại.\" kèm nút \"Thử lại\" (CMR-07).")

tc("[Error] Kiểm tra lỗi 500 (Chi tiết) → không có Thử lại",
   "Simulate API chi tiết trả 500.",
   "1. Tap 1 card.",
   "1. Hiển thị \"Hệ thống đang bận. Vui lòng thử lại sau.\" — KHÔNG có nút Thử lại.")

tc("[Error] Kiểm tra HTTP 404 lịch hẹn không tồn tại → toast + back về Danh sách",
   "Simulate API chi tiết trả HTTP 404 (record bị xóa).",
   "1. Tap 1 card.",
   "1. Hiển thị \"Nội dung không tồn tại hoặc đã bị xóa.\" → app tự quay về Danh sách.")

tc("[Error] Kiểm tra timeout > 10s (Chi tiết) → message + Thử lại",
   "Simulate API chi tiết mất > 10s.",
   "1. Tap 1 card.\n2. Chờ qua 10s.",
   "2. Hiển thị \"Yêu cầu đã hết thời gian chờ. Vui lòng thử lại.\" kèm nút \"Thử lại\" (CMR-16).")

tc("[Error 401] Kiểm tra 401 + refresh expired trên Chi tiết → redirect login",
   "Đang ở Chi tiết; simulate token expired.",
   "1. Pull-to-refresh trên Chi tiết.",
   "1. Toast \"Phiên đăng nhập hết hạn.\"; redirect về màn [Đăng nhập] (AC16).")

tc("[i18n Chi tiết] Kiểm tra text cứng label 5 section đổi theo ngôn ngữ",
   "Đang ở Chi tiết ngôn ngữ VI.",
   "1. Đổi ngôn ngữ sang EN.\n2. Quay lại Chi tiết.",
   "2.\n- Label các field (\"Mã thủ tục\", \"Tên thủ tục\", \"Lĩnh vực\", ...) đổi sang EN.\n- Text động từ API (giá trị Tên thủ tục, Lĩnh vực, Ghi chú, v.v.) giữ nguyên ngôn ngữ gốc (AC20, CMR-17).")

tc("[A11y] Kiểm tra Screen reader đọc được 5 section trong Chi tiết",
   "Bật screen reader.",
   "1. Mở Chi tiết.\n2. Vuốt/điều hướng qua 5 section.",
   "2. Mỗi field (label + giá trị) được đọc đầy đủ (AC21).")

# --- Check common (Section 2) ---
check_type("Check common")

sub_label("Kiểm tra các trường hợp phổ biến UI/UX")

tc("Kiểm tra hiển thị dữ liệu tối đa (maxlength)",
   "",
   "1. Kiểm tra hiển thị dữ liệu tối đa (maxlength)",
   "1. Hiển thị đúng độ dài tối đa")

tc("Kiểm tra khôi phục/phóng to/thu nhỏ ứng dụng",
   "",
   "1. Thực hiện khôi phục, phóng to, thu nhỏ ứng dụng",
   "1. Không xảy ra lỗi bất thường")

tc("Kiểm tra tính nhất quán của các thông báo",
   "",
   "1. Kiểm tra tính nhất quán của các thông báo",
   "1. Xác nhận thông báo lỗi:\n- Thông báo lỗi inline: hiển thị dưới ô nhập liệu bị lỗi, màu đỏ\n- Thông báo lỗi dạng toast: hiển thị ở giữa hoặc phía dưới màn hình")

tc("Kiểm tra hiển thị khi thiết bị ở chế độ dọc (thiết bị xoay theo chiều đứng)",
   "",
   "1. Kiểm tra hiển thị khi thiết bị ở chế độ dọc",
   "1. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị ở chế độ ngang (thiết bị xoay theo chiều ngang)",
   "",
   "1. Kiểm tra hiển thị khi thiết bị ở chế độ ngang",
   "1. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi chuyển đổi giữa chế độ dọc và ngang",
   "",
   "1. Kiểm tra hiển thị khi chuyển từ chế độ dọc sang ngang\n2. Kiểm tra hiển thị khi chuyển từ chế độ ngang sang dọc",
   "1 & 2. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ lớn nhất",
   "",
   "1. Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ lớn nhất",
   "1. Giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ nhỏ nhất",
   "",
   "1. Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ nhỏ nhất",
   "1. Giao diện không bị vỡ")


sub_label("Kiểm tra tương tác cơ bản với thiết bị")

tc("Xác nhận hiển thị của ứng dụng khi người dùng chạm vào nút [Quay lại] trên thiết bị Android",
   "",
   "1. Mở ứng dụng\n2. Người dùng chạm vào nút [Quay lại] trên thiết bị Android\n3. Xác nhận hiển thị",
   "3. Quay lại màn hình trước đó")

tc("Xác nhận hiển thị của ứng dụng khi người dùng vuốt từ trái sang phải trên thiết bị iOS",
   "",
   "1. Mở ứng dụng\n2. Người dùng vuốt từ trái sang phải trên thiết bị iOS\n3. Xác nhận hiển thị",
   "3. Quay lại màn hình trước đó")

tc("Xác nhận hiển thị của ứng dụng khi người dùng tắt và mở lại ứng dụng",
   "",
   "1. Mở ứng dụng\n2. Người dùng tắt ứng dụng\n3. Mở lại ứng dụng\n4. Xác nhận hiển thị",
   "4. Ứng dụng mở lại từ trạng thái ban đầu")

tc("Kiểm tra chế độ đa nhiệm (multitasking)",
   "",
   "1. Mở ứng dụng\n2. Trở về màn hình chính (không tắt ứng dụng)\n3. Mở lại ứng dụng\n4. Xác nhận hiển thị",
   "4. Ứng dụng giữ nguyên ở trạng thái hiện tại")

tc("Xác nhận hiển thị của ứng dụng khi người dùng khóa và mở khóa màn hình thiết bị",
   "",
   "1. Mở ứng dụng\n2. Khóa thiết bị\n3. Mở khóa thiết bị\n4. Xác nhận hiển thị",
   "4. Giữ nguyên trạng thái hiện tại của ứng dụng")

tc("Kiểm tra hành động kéo xuống để làm mới (pull-to-refresh)",
   "Tiền điều kiện: Màn hình hỗ trợ tính năng kéo để làm mới",
   "1. Người dùng ở màn Chi tiết\n2. Kéo xuống để làm mới",
   "2. Hiển thị dữ liệu chi tiết mới nhất")

tc("Kiểm tra hành động cuộn xuống để tải thêm (scroll-down-to-load-more)",
   "N/A - Màn Chi tiết không hỗ trợ lazy load (chỉ Danh sách hỗ trợ).",
   "N/A",
   "N/A",
   "N/A")

tc("Kiểm tra phản hồi của ứng dụng khi thiết bị nhận được thông báo từ ứng dụng khác",
   "Tiền điều kiện: Ứng dụng khác được phép gửi thông báo",
   "1. Mở Chi tiết\n2. Ứng dụng khác gửi thông báo\n3. Xác nhận hiển thị",
   "3. Không có lỗi nào xảy ra")

# ============================================================
# Column widths
# ============================================================
col_widths = {
    'A': 14.25, 'B': 39.0, 'C': 37.38, 'D': 50.38, 'E': 48.63,
    'F': 16.0, 'G': 16.0, 'K': 16.0, 'O': 16.0, 'Q': 16.0,
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

for r in range(1, 8):
    ws.row_dimensions[r].height = 15.75

wb.save(output_path)
print(f"DONE. Total TC: {tc_counter[0]}. Saved to: {output_path}")
