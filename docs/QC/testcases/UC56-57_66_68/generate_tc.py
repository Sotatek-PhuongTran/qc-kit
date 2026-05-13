import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

base_dir = r"c:\Users\admin\Desktop\CMC\mbfs-bqa\docs\QC-MOBILE\testcases\UC56-57_66_68"
xlsx_path = f"{base_dir}\\UC56-57_66_68_tin-tuc_testcases_20260511_v1.xlsx"

wb = load_workbook(xlsx_path)
ws = wb["Tin tức"]

# Unmerge all cells and clear data from row 8 onwards
for merge in list(ws.merged_cells.ranges):
    if merge.min_row >= 8:
        ws.unmerge_cells(str(merge))
if ws.max_row >= 8:
    ws.delete_rows(8, ws.max_row - 7)

# Load all data
all_data = []
for f in ["tc_data.json", "tc_data2.json", "tc_data3.json"]:
    with open(f"{base_dir}\\{f}", "r", encoding="utf-8") as fh:
        all_data.extend(json.load(fh))

# Styles
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
section_fill = PatternFill(start_color="FFA4C2F4", end_color="FFA4C2F4", fill_type="solid")
check_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
sub_fill = PatternFill(start_color="FFE3FAFD", end_color="FFE3FAFD", fill_type="solid")

font_section = Font(name="Times New Roman", size=11, bold=True)
font_check = Font(name="Times New Roman", size=11, bold=True)
font_sub = Font(name="Times New Roman", size=11, bold=True, italic=True)
font_tc = Font(name="Times New Roman", size=11)

align_wrap = Alignment(wrap_text=True, vertical="top")
align_center = Alignment(wrap_text=True, vertical="center", horizontal="center")

row = 8
tc_count = 0

for entry in all_data:
    entry_type = entry[0]
    if entry_type == "section":
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            cell.fill = section_fill
            cell.border = thin_border
            cell.font = font_section
        ws.cell(row=row, column=1, value=entry[1])
        ws.cell(row=row, column=1).alignment = align_wrap
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        row += 1
    elif entry_type == "check":
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            cell.fill = check_fill
            cell.border = thin_border
            cell.font = font_check
        ws.cell(row=row, column=1, value=entry[1])
        ws.cell(row=row, column=1).alignment = align_wrap
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        row += 1
    elif entry_type == "sub":
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            cell.fill = sub_fill
            cell.border = thin_border
            cell.font = font_sub
        ws.cell(row=row, column=1, value=entry[1])
        ws.cell(row=row, column=1).alignment = align_wrap
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        row += 1
    elif entry_type == "tc":
        tc_id, title, precond, steps, expected, note = entry[1], entry[2], entry[3], entry[4], entry[5], entry[6]
        tc_count += 1
        # Col A: TC ID
        ws.cell(row=row, column=1, value=tc_id).font = font_tc
        ws.cell(row=row, column=1).alignment = align_center
        # Col B: Title
        ws.cell(row=row, column=2, value=title).font = font_tc
        ws.cell(row=row, column=2).alignment = align_wrap
        # Col C: Pre-condition
        ws.cell(row=row, column=3, value=precond).font = font_tc
        ws.cell(row=row, column=3).alignment = align_wrap
        # Col D: Steps
        ws.cell(row=row, column=4, value=steps).font = font_tc
        ws.cell(row=row, column=4).alignment = align_wrap
        # Col E: Expected Result
        ws.cell(row=row, column=5, value=expected).font = font_tc
        ws.cell(row=row, column=5).alignment = align_wrap
        # Col F: Note
        ws.cell(row=row, column=6, value=note).font = font_tc
        ws.cell(row=row, column=6).alignment = align_wrap
        # Border all cols
        for col in range(1, 18):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

# Column widths
col_widths = {'A': 14.25, 'B': 39.0, 'C': 37.38, 'D': 50.38, 'E': 48.63, 'F': 16.0, 'G': 16.0, 'K': 16.0, 'O': 16.0, 'Q': 16.0}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

wb.save(xlsx_path)
print(f"Done! Wrote {tc_count} test cases, total {row - 8} rows (including section/check/sub headers).")