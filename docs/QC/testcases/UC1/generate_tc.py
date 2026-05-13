import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

template_path = r"c:\Users\admin\Desktop\CMC\mbfs-bqa\.claude\skills\qc-tc-design-MOBILE\templates\[MBFS] Template TestCase - Mobile.xlsx"
output_path = r"c:\Users\admin\Desktop\CMC\mbfs-bqa\docs\QC-MOBILE\testcases\UC1\UC1_trang-chu-dashboard_testcases_20260507_v1.xlsx"

shutil.copy2(template_path, output_path)
wb = load_workbook(output_path)

ws = wb["Testcase"]
ws.title = "Trang chu Dashboard"

font_default = Font(name='Calibri', size=11, color='FF000000')
font_bold = Font(name='Calibri', size=11, color='FF000000', bold=True)
align_lw = Alignment(horizontal='left', vertical='center', wrap_text=True)
fill_section = PatternFill(start_color='FFA4C2F4', end_color='FFA4C2F4', fill_type='solid')
fill_check = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
fill_sub = PatternFill(start_color='FFE3FAFD', end_color='FFE3FAFD', fill_type='solid')

ws.cell(row=1, column=2).value = "Cổng Một Cửa Đầu Tư Quốc Gia"
ws.cell(row=1, column=4).value = "MBFS_UC-01_TrangChuDashboard_Testcase_v1"
ws.cell(row=3, column=4).value = "Màn hình Trang chủ (Dashboard) Mobile"
ws.cell(row=4, column=2).value = "QC Agent"
ws.cell(row=4, column=4).value = "2026-05-07"

for merge in list(ws.merged_cells.ranges):
    if merge.min_row >= 8:
        ws.unmerge_cells(str(merge))
for row in range(8, ws.max_row + 1):
    for col in range(1, 19):
        ws.cell(row=row, column=col).value = None
        ws.cell(row=row, column=col).fill = PatternFill()

cur = [8]

def section(title):
    r = cur[0]
    for c in range(1, 19):
        ws.cell(row=r, column=c).fill = fill_section
    ws.cell(row=r, column=2).value = title
    ws.cell(row=r, column=2).font = font_bold
    ws.cell(row=r, column=2).alignment = align_lw
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
    cur[0] += 1

def check_type(title):
    r = cur[0]
    for c in range(1, 19):
        ws.cell(row=r, column=c).fill = fill_check
    ws.cell(row=r, column=2).value = title
    ws.cell(row=r, column=2).font = font_bold
    ws.cell(row=r, column=2).alignment = align_lw
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
    cur[0] += 1

def sub_label(title):
    r = cur[0]
    for c in range(1, 19):
        ws.cell(row=r, column=c).fill = fill_sub
    ws.cell(row=r, column=2).value = title
    ws.cell(row=r, column=2).font = font_bold
    ws.cell(row=r, column=2).alignment = align_lw
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
    cur[0] += 1

tc_counter = [0]
def tc(title, pre, steps, expected, note=""):
    r = cur[0]
    tc_counter[0] += 1
    ws.cell(row=r, column=1).value = f"TC_{tc_counter[0]:03d}"
    ws.cell(row=r, column=1).font = font_default
    ws.cell(row=r, column=1).alignment = align_lw
    ws.cell(row=r, column=2).value = title
    ws.cell(row=r, column=2).font = font_default
    ws.cell(row=r, column=2).alignment = align_lw
    ws.cell(row=r, column=3).value = pre
    ws.cell(row=r, column=3).font = font_default
    ws.cell(row=r, column=3).alignment = align_lw
    ws.cell(row=r, column=4).value = steps
    ws.cell(row=r, column=4).font = font_default
    ws.cell(row=r, column=4).alignment = align_lw
    ws.cell(row=r, column=5).value = expected
    ws.cell(row=r, column=5).font = font_default
    ws.cell(row=r, column=5).alignment = align_lw
    ws.cell(row=r, column=6).value = note
    ws.cell(row=r, column=6).font = font_default
    ws.cell(row=r, column=6).alignment = align_lw
    ws.cell(row=r, column=7).value = "Untested"
    ws.cell(row=r, column=11).value = "Untested"
    cur[0] += 1

# ============================================================
# SECTION: 1. Màn hình Trang chủ (Dashboard)
# ============================================================
section("1. Màn hình Trang chủ (Dashboard)")

# --- Check UI/UX ---
check_type("Check UI/UX")

tc("Kiểm tra UI/UX màn hình [Trang chủ] khi có dữ liệu",
   "Đăng nhập thành công vào ứng dụng",
   "1. Kiểm tra hiển thị màn hình [Trang chủ]",
   '1.\n- Hiển thị đầy đủ các item, màu sắc, font chữ, layout giống Design\n(Tham khảo ảnh 1. Trang chủ sheet WFDesign)\n- Header: Hamburger (☰), Logo + "Cổng Đầu Tư", Icon Ngôn ngữ "VI", Icon Thông báo (🔔), Icon Người dùng\n- Card thông tin: Avatar (icon mặc định), Tên đầy đủ, Vai trò\n- Quick Access: 6 mục cố định:\n  + Hướng dẫn sử dụng\n  + Quản lý hồ sơ\n  + Quản lý đặt lịch\n  + Khu công nghiệp / KKT\n  + Câu hỏi (FAQ)\n  + Văn bản pháp luật\n- Tin tức: Tiêu đề "Tin tức", nút "Xem tất cả" (chữ đỏ), danh sách tối đa 5 tin cuộn ngang\n- Chatbot: Icon nổi góc dưới phải\n- Footer: 4 tab (Trang chủ active đỏ, Thủ tục, KCN/KKT, Cài đặt)')

tc("Kiểm tra UI/UX màn hình [Trang chủ] khi tin tức rỗng",
   "Đăng nhập thành công, hệ thống không có tin tức",
   "1. Kiểm tra hiển thị màn hình [Trang chủ]",
   '1.\n- Header, Card thông tin, Quick Access, Chatbot, Footer: hiển thị bình thường\n- Vùng Tin tức: hiển thị "Không có dữ liệu." căn giữa vùng nội dung (CMR-14)')

tc("Kiểm tra hiển thị Loading state khi tải dữ liệu",
   "Đăng nhập thành công, đang chờ API phản hồi",
   "1. Kiểm tra hiển thị màn hình [Trang chủ] trong khi chờ API",
   "1.\n- Card thông tin: skeleton loading/spinner\n- Tin tức: skeleton loading\n- Quick Access: hiển thị cố định (không loading)\n- Footer: hiển thị cố định")

tc("[Card thông tin] Kiểm tra hiển thị tên đầy đủ quá dài (truncate)",
   "Đăng nhập với tài khoản có tên dài vượt quá 1 dòng hiển thị",
   "1. Kiểm tra hiển thị trường Tên đầy đủ trên Card thông tin",
   '1. Tên hiển thị tối đa 1 dòng, phần vượt quá bị cắt và hiển thị dấu "..." ở cuối',
   "BVA")

tc("[Card thông tin] Kiểm tra hiển thị vai trò \"Nhà đầu tư Việt Nam\"",
   "Đăng nhập với tài khoản role Nhà đầu tư Việt Nam",
   "1. Kiểm tra hiển thị trường Vai trò trên Card thông tin",
   '1. Hiển thị text "Nhà đầu tư Việt Nam"')

tc("[Card thông tin] Kiểm tra hiển thị vai trò \"Nhà đầu tư nước ngoài\"",
   "Đăng nhập với tài khoản role Nhà đầu tư nước ngoài",
   "1. Kiểm tra hiển thị trường Vai trò trên Card thông tin",
   '1. Hiển thị text "Nhà đầu tư nước ngoài"')

tc("[Card thông tin] Kiểm tra hiển thị vai trò \"Tổ chức/Doanh nghiệp\"",
   "Đăng nhập với tài khoản role Tổ chức/Doanh nghiệp",
   "1. Kiểm tra hiển thị trường Vai trò trên Card thông tin",
   '1. Hiển thị text "Tổ chức/Doanh nghiệp"')

tc("[Avatar] Kiểm tra Avatar luôn hiển thị icon mặc định",
   "Đăng nhập thành công",
   "1. Kiểm tra hiển thị Avatar trên Card thông tin",
   "1. Avatar hiển thị icon mặc định hệ thống (hình tròn)\n- Không lấy ảnh từ profile người dùng\n- Không tap được")

tc("[Tin tức] Kiểm tra card tin tức hiển thị đầy đủ thông tin",
   "Đăng nhập thành công, có tin tức",
   "1. Kiểm tra hiển thị 1 card tin tức",
   "1. Mỗi card hiển thị đầy đủ:\n- Ảnh đại diện (thumbnail bo góc)\n- Nhãn phân loại (category tag)\n- Ngày đăng (DD/MM/YYYY)\n- Tiêu đề bài viết (tối đa 2 dòng)")

tc("[Tin tức] Kiểm tra category tag hiển thị đúng màu sắc",
   "Đăng nhập thành công, có tin tức các loại",
   "1. Kiểm tra hiển thị nhãn phân loại trên card tin tức",
   '1. Nhãn phân loại hiển thị đúng màu:\n- Đỏ: "Chính sách"\n- Xanh: "Tin đầu tư"\n- Cam: "Thành công"',
   "EP: 3 partitions")

tc("[Tin tức] Kiểm tra tiêu đề bài viết quá dài (truncate 2 dòng)",
   "Đăng nhập thành công, có tin tức với tiêu đề rất dài",
   "1. Kiểm tra hiển thị tiêu đề bài viết trên card tin tức",
   '1. Tiêu đề hiển thị tối đa 2 dòng, phần vượt quá bị cắt và hiển thị dấu "..."',
   "BVA")

tc("[Icon Thông báo] Kiểm tra hiển thị red dot khi có thông báo chưa đọc",
   "Đăng nhập thành công, có thông báo chưa đọc",
   "1. Kiểm tra hiển thị icon Thông báo (🔔) trên Header",
   "1. Icon 🔔 hiển thị red dot (chấm đỏ) ở góc phải trên\n- Không kèm số")

tc("[Icon Thông báo] Kiểm tra không hiển thị red dot khi đã đọc hết",
   "Đăng nhập thành công, tất cả thông báo đã đọc",
   "1. Kiểm tra hiển thị icon Thông báo (🔔) trên Header",
   "1. Icon 🔔 hiển thị bình thường, không có red dot")

tc("[Footer] Kiểm tra tab Trang chủ ở trạng thái Active",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Kiểm tra hiển thị Footer",
   '1. Tab "Trang chủ": icon + text màu đỏ (Active)\n- 3 tab còn lại: trạng thái Inactive')

# --- Check Function ---
check_type("Check Function")

tc("[Nút Hamburger] Kiểm tra mở Sidebar Navigation",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Nhấp vào nút [Hamburger ☰] góc trái header",
   "1. Mở Sidebar Navigation từ bên trái màn hình.\n- Sidebar chứa toàn bộ menu điều hướng chính của ứng dụng. Danh sách menu: Trang chủ, Giới thiệu, Lĩnh vực đầu tư, Khu vực đầu tư, Liên hệ, Thủ tục hành chính, Quản lý hồ sơ, Quản lý đặt lịch, Phản ánh kiến nghị, Cấu hình tài khoản.\n- Vùng bên ngoài Sidebar hiển thị mờ")

tc("[Sidebar] Kiểm tra tap item điều hướng đúng",
   "Đăng nhập thành công, Sidebar đang mở",
   '1. Nhấp vào item "Quản lý hồ sơ" trong Sidebar',
   "1. Đóng Sidebar\n2. Điều hướng đến màn hình Quản lý hồ sơ (UC45)")

tc('[Sidebar] Kiểm tra tap "Trang chủ" điều hướng đúng',
   "Đăng nhập thành công, Sidebar đang mở, đang ở màn hình khác (không phải Trang chủ)",
   '1. Nhấp vào item "Trang chủ" trong Sidebar',
   "1. Đóng Sidebar\n2. Điều hướng về màn hình Trang chủ (UC1)")

tc('[Sidebar] Kiểm tra tap "Quản lý đặt lịch" điều hướng đúng',
   "Đăng nhập thành công, Sidebar đang mở",
   '1. Nhấp vào item "Quản lý đặt lịch" trong Sidebar',
   "1. Đóng Sidebar\n2. Điều hướng đến màn hình Quản lý đặt lịch (UC42)")

tc('[Sidebar] Kiểm tra tap "Phản ánh kiến nghị" điều hướng đúng',
   "Đăng nhập thành công, Sidebar đang mở",
   '1. Nhấp vào item "Phản ánh kiến nghị" trong Sidebar',
   "1. Đóng Sidebar\n2. Điều hướng đến màn hình Phản ánh kiến nghị")

tc('[Sidebar] Kiểm tra tap "Cấu hình tài khoản" điều hướng đúng',
   "Đăng nhập thành công, Sidebar đang mở",
   '1. Nhấp vào item "Cấu hình tài khoản" trong Sidebar',
   "1. Đóng Sidebar\n2. Điều hướng đến màn hình Cấu hình tài khoản (UC249)")

tc("[Sidebar] Kiểm tra đóng bằng tap vùng mờ bên ngoài",
   "Đăng nhập thành công, Sidebar đang mở",
   "1. Nhấp vào vùng mờ bên ngoài Sidebar",
   "1. Đóng Sidebar\n- Không điều hướng đến bất kỳ màn hình nào")

tc("[Icon Ngôn ngữ] Kiểm tra mở popup chọn ngôn ngữ",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Nhấp vào icon Ngôn ngữ trên Header",
   '1. Mở popup "Chọn ngôn ngữ"\n- Hiển thị 5 ngôn ngữ: Tiếng Việt, English, 中文, 日本語, 한국어\n- Ngôn ngữ đang chọn được đánh dấu check đỏ')

tc("[Popup Ngôn ngữ] Kiểm tra chuyển sang English (EN)",
   "Đăng nhập thành công, popup Ngôn ngữ đang mở, ngôn ngữ hiện tại VI",
   '1. Nhấp vào "English" trong popup',
   '1. Đóng popup\n2. Mã ngôn ngữ trên Header cập nhật thành "EN"\n3. Áp dụng ngôn ngữ cho toàn hệ thống\n4. Tin tức tải lại theo ngôn ngữ English (không dùng cache)')

tc("[Popup Ngôn ngữ] Kiểm tra chọn ngôn ngữ hiện tại (không thay đổi)",
   "Đăng nhập thành công, popup Ngôn ngữ đang mở, ngôn ngữ hiện tại VI",
   '1. Nhấp vào "Tiếng Việt" (ngôn ngữ đang active)',
   '1. Đóng popup\n- Không thay đổi ngôn ngữ hệ thống\n- Mã header vẫn là "VI"\n- Tin tức không reload')

tc("[Popup Ngôn ngữ] Kiểm tra chuyển sang tiếng Nhật (JA)",
   "Đăng nhập thành công, popup Ngôn ngữ đang mở",
   '1. Nhấp vào "日本語" trong popup',
   '1. Đóng popup\n2. Mã ngôn ngữ Header cập nhật thành "JA"\n3. Tin tức tải lại theo ngôn ngữ Nhật')

tc("[Popup Ngôn ngữ] Kiểm tra chuyển sang tiếng Hàn (KO)",
   "Đăng nhập thành công, popup Ngôn ngữ đang mở",
   '1. Nhấp vào "한국어" trong popup',
   '1. Đóng popup\n2. Mã ngôn ngữ Header cập nhật thành "KO"\n3. Tin tức tải lại theo ngôn ngữ Hàn')

tc("[Popup Ngôn ngữ] Kiểm tra chuyển sang tiếng Trung (ZH)",
   "Đăng nhập thành công, popup Ngôn ngữ đang mở",
   '1. Nhấp vào "中文" trong popup',
   '1. Đóng popup\n2. Mã ngôn ngữ Header cập nhật thành "ZH"\n3. Tin tức tải lại theo ngôn ngữ Trung')

tc("[Ngôn ngữ] Kiểm tra ngôn ngữ được lưu lại sau khi đóng và mở lại ứng dụng",
   "Đăng nhập thành công, đã chuyển ngôn ngữ sang EN",
   "1. Đóng ứng dụng hoàn toàn\n2. Mở lại ứng dụng",
   '2. Ngôn ngữ vẫn là "EN"\n- Mã header hiển thị "EN"\n- Tin tức hiển thị tiếng Anh')

tc("[Icon Thông báo] Kiểm tra tap điều hướng đến màn hình Thông báo",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Nhấp vào icon Thông báo (🔔) trên Header",
   "1. Điều hướng đến màn hình Thông báo")

tc("[Icon Người dùng] Kiểm tra tap điều hướng đến màn hình Cấu hình tài khoản",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Nhấp vào icon Người dùng góc phải Header",
   "1. Điều hướng đến màn hình Cấu hình tài khoản")

tc('[Quick Access] Kiểm tra tap "Hướng dẫn sử dụng"',
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào card "Hướng dẫn sử dụng" trong Quick Access',
   "1. Điều hướng đến màn hình Hướng dẫn sử dụng")

tc('[Quick Access] Kiểm tra tap "Quản lý hồ sơ"',
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào card "Quản lý hồ sơ" trong Quick Access',
   "1. Điều hướng đến màn hình Quản lý hồ sơ")

tc('[Quick Access] Kiểm tra tap "Quản lý đặt lịch"',
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào card "Quản lý đặt lịch" trong Quick Access',
   "1. Điều hướng đến màn hình Quản lý đặt lịch")

tc('[Quick Access] Kiểm tra tap "Khu công nghiệp / KKT"',
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào card "Khu công nghiệp / KKT" trong Quick Access',
   "1. Điều hướng đến màn hình Tra cứu KCN/KKT")

tc('[Quick Access] Kiểm tra tap "Câu hỏi (FAQ)"',
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào card "Câu hỏi (FAQ)" trong Quick Access',
   "1. Điều hướng đến màn hình Câu hỏi thường gặp")

tc('[Quick Access] Kiểm tra tap "Văn bản pháp luật"',
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào card "Văn bản pháp luật" trong Quick Access',
   "1. Điều hướng đến màn hình Văn bản pháp luật")

tc("[Tin tức] Kiểm tra tap card → chi tiết bài viết",
   "Đăng nhập thành công, có tin tức hiển thị",
   "1. Nhấp vào 1 card tin tức trong danh sách",
   "1. Điều hướng đến màn hình chi tiết bài viết (UC55-68)")

tc('[Nút "Xem tất cả"] Kiểm tra tap → danh sách tin tức',
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào nút "Xem tất cả" bên phải tiêu đề "Tin tức"',
   "1. Điều hướng đến màn hình danh sách Tin tức đầy đủ (UC55-68)")

tc("[Tin tức] Kiểm tra cuộn ngang (horizontal scroll)",
   "Đăng nhập thành công, có >= 3 tin tức",
   "1. Vuốt ngang (swipe left) trên danh sách tin tức",
   "1. Danh sách tin tức cuộn ngang, hiển thị các tin tiếp theo")

tc("[Tin tức] Kiểm tra hiển thị tối đa 5 bài, sắp xếp giảm dần",
   "Đăng nhập thành công, hệ thống có > 5 tin tức",
   "1. Kiểm tra số lượng tin tức hiển thị trên Trang chủ",
   "1. Hiển thị tối đa 5 tin\n- Sắp xếp theo thời gian đăng giảm dần (mới nhất trước)")

tc("[Tin tức] Kiểm tra khi có ít hơn 5 tin",
   "Đăng nhập thành công, hệ thống chỉ có 2 tin tức",
   "1. Kiểm tra hiển thị danh sách tin tức",
   "1. Hiển thị đúng 2 tin có sẵn\n- Không hiển thị placeholder hay lỗi cho vị trí trống",
   "BVA: < 5 tin")

tc("[Icon Chatbot] Kiểm tra tap mở Chatbot (UC60)",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Nhấp vào icon Chatbot (góc dưới phải)",
   "1. Điều hướng đến màn hình Chatbot (UC60)")

tc("[Icon Chatbot] Kiểm tra Drag & Drop di chuyển vị trí",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Chạm giữ icon Chatbot\n2. Kéo thả đến vị trí khác trên màn hình",
   "2. Icon Chatbot di chuyển đến vị trí mới")

tc("[Icon Chatbot] Kiểm tra reset vị trí khi quay lại Trang chủ",
   "Đăng nhập thành công, đã kéo Chatbot sang vị trí khác",
   "1. Điều hướng sang màn hình khác (ví dụ: UC45)\n2. Quay lại Trang chủ",
   "2. Icon Chatbot tự động trở về vị trí mặc định (góc dưới phải)")

tc("[Footer] Kiểm tra tap tab Trang chủ refresh dữ liệu",
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào tab "Trang chủ" trên Footer',
   "1. Dữ liệu trên màn hình được refresh (tải lại từ đầu)")

tc("[Footer] Kiểm tra tap tab Thủ tục",
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào tab "Thủ tục" trên Footer',
   "1. Điều hướng đến màn hình Tra cứu Văn bản & Thủ tục")

tc("[Footer] Kiểm tra tap tab KCN/KKT",
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào tab "KCN/KKT" trên Footer',
   "1. Điều hướng đến màn hình Tra cứu KCN/KKT")

tc("[Footer] Kiểm tra tap tab Cài đặt",
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp vào tab "Cài đặt" trên Footer',
   "1. Điều hướng đến màn hình Cấu hình tài khoản")

tc("[Thông báo] Kiểm tra polling 30s cập nhật red dot",
   "Đăng nhập thành công, đang ở Trang chủ, ban đầu không có thông báo chưa đọc",
   "1. Chờ hệ thống gửi thông báo mới\n2. Chờ tối đa 30 giây\n3. Kiểm tra icon Thông báo",
   "3. Red dot xuất hiện trên icon 🔔 mà không cần rời đi và quay lại")

tc("[Pull to Refresh] Kiểm tra kéo xuống refresh thành công",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Kéo xuống từ đầu màn hình",
   "1. Hiển thị spinner ở đầu danh sách\n2. Dữ liệu được tải lại từ đầu\n3. Spinner ẩn sau khi refresh thành công",
   "CMR-13")

tc("[Pull to Refresh] Kiểm tra refresh thất bại giữ dữ liệu cũ",
   "Đăng nhập thành công, đang ở Trang chủ, tắt mạng",
   "1. Kéo xuống từ đầu màn hình",
   '1. Hiển thị spinner\n2. Hiển thị thông báo lỗi theo CMR-07:\n- Lỗi mạng: Hiển thị thông báo "Không thể kết nối. Vui lòng kiểm tra mạng và thử lại." kèm nút "Thử lại"\n3. Giữ nguyên dữ liệu cũ trên màn hình',
   "CMR-13")

tc("[Pull to Refresh] Kiểm tra không duplicate API khi đang loading",
   "Đăng nhập thành công, đang pull to refresh (spinner hiển thị)",
   "1. Kéo xuống lần nữa trong khi đang loading",
   "1. Không trigger lại API tương tự\n- Chỉ 1 request được gửi",
   "CMR-13")

tc("[Lỗi mạng] Kiểm tra hiển thị thông báo lỗi mạng",
   "Đăng nhập thành công, tắt kết nối mạng",
   "1. Mở ứng dụng (hoặc refresh Trang chủ)",
   '1. Hiển thị thông báo "Không thể kết nối. Vui lòng kiểm tra mạng và thử lại." kèm nút "Thử lại" tại section tương ứng',
   "CMR-07")

tc("[Lỗi API 500] Kiểm tra hiển thị thông báo hệ thống bận",
   "Đăng nhập thành công, API trả về HTTP 500",
   "1. Mở Trang chủ",
   '1. Hiển thị thông báo "Hệ thống đang bận. Vui lòng thử lại sau." tại section tương ứng',
   "CMR-07")

tc("[Timeout] Kiểm tra hiển thị thông báo timeout > 10 giây",
   "Đăng nhập thành công, API phản hồi chậm > 10 giây",
   "1. Mở Trang chủ\n2. Chờ quá 10 giây",
   '2. Hiển thị thông báo "Yêu cầu đã hết thời gian chờ. Vui lòng thử lại." kèm nút "Thử lại"',
   "CMR-16")

tc("[Timeout BVA] Kiểm tra API phản hồi đúng 10 giây",
   "Đăng nhập thành công, API phản hồi đúng 10 giây",
   "1. Mở Trang chủ\n2. API phản hồi tại mốc 10 giây",
   "2. Hiển thị dữ liệu thành công\n- Không trigger timeout",
   "BVA: boundary")

tc("Kiểm tra hiển thị data ở Card thông tin user khi API user info bị failed",
   "Đăng nhập thành công, API user info lỗi, API tin tức OK",
   "1. Mở Trang chủ",
   "1. Section Card thông tin: hiển thị thông báo lỗi theo CMR-07\n- Section Tin tức: hiển thị bình thường\n- Quick Access, Footer: hiển thị bình thường")

tc("Kiểm tra hiển thị data ở Card tin tức khi API tin tức bị failed",
   "Đăng nhập thành công, API tin tức lỗi, API user info OK",
   "1. Mở Trang chủ",
   "1. Section Card thông tin: hiển thị bình thường\n- Section Tin tức: hiển thị thông báo lỗi\n- Quick Access, Footer: hiển thị bình thường")

tc("Kiểm tra hiển thị data ở Card thông tin user, card tin tức khi tất cả API fails",
   "Đăng nhập thành công, tất cả API trả về lỗi",
   "1. Mở Trang chủ",
   "1. Mỗi section hiển thị thông báo lỗi riêng theo CMR-07\n- Quick Access vẫn hiển thị (cố định)\n- Footer vẫn hiển thị")

tc('[Nút Thử lại] Kiểm tra tap retry sau lỗi mạng',
   "Đăng nhập thành công, đang hiển thị lỗi mạng, đã bật lại mạng",
   '1. Nhấp vào nút "Thử lại"',
   "1. Gọi lại API\n2. Hiển thị loading indicator\n3. Hiển thị dữ liệu thành công")

tc("[Debounce] Kiểm tra double tap Quick Access không duplicate",
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp nhanh liên tục 2 lần vào card "Quản lý hồ sơ"',
   "1. Chỉ điều hướng 1 lần đến UC45\n- Không push duplicate screen")

tc("[Debounce] Kiểm tra double tap Footer tab không duplicate",
   "Đăng nhập thành công, đang ở Trang chủ",
   '1. Nhấp nhanh liên tục 2 lần vào tab "Thủ tục"',
   "1. Chỉ điều hướng 1 lần đến UC69/UC73\n- Không push duplicate screen")

tc("[App lifecycle] Kiểm tra force close giữ session",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Force close ứng dụng\n2. Mở lại ứng dụng",
   "2. Quay về Trang chủ\n- Giữ nguyên session đăng nhập\n- Không yêu cầu đăng nhập lại")

tc("[App lifecycle] Kiểm tra uninstall yêu cầu đăng nhập lại",
   "Đã đăng nhập trước đó",
   "1. Xóa ứng dụng (uninstall)\n2. Cài đặt lại ứng dụng\n3. Mở ứng dụng",
   "3. Yêu cầu đăng nhập lại từ đầu\n- Không restore session cũ")

tc("[Quay lại Trang chủ] Kiểm tra refresh khi quay lại từ màn hình khác",
   "Đăng nhập thành công, đã điều hướng sang UC45",
   '1. Nhấp vào tab "Trang chủ" trên Footer để quay lại',
   "1. Dữ liệu Trang chủ được refresh\n- Hiển thị thông tin mới nhất")

# --- Check common ---
check_type("Check common")
sub_label("Kiểm tra các trường hợp phổ biến UI/UX")

tc("Kiểm tra hiển thị dữ liệu tối đa (maxlength)",
   "Đăng nhập thành công",
   "1. Kiểm tra hiển thị dữ liệu tối đa (maxlength)",
   "1. Hiển thị đúng độ dài tối đa")

tc("Kiểm tra khôi phục/phóng to/thu nhỏ ứng dụng",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Thực hiện khôi phục, phóng to, thu nhỏ ứng dụng",
   "1. Không xảy ra lỗi bất thường")

tc("Kiểm tra tính nhất quán của các thông báo",
   "Đăng nhập thành công",
   "1. Kiểm tra tính nhất quán của các thông báo",
   "1. Xác nhận thông báo lỗi:\n- Thông báo lỗi inline: hiển thị dưới ô nhập liệu bị lỗi, màu đỏ\n- Thông báo lỗi dạng toast: hiển thị ở giữa hoặc phía dưới màn hình")

tc("Kiểm tra hiển thị khi thiết bị ở chế độ dọc",
   "Đăng nhập thành công",
   "1. Kiểm tra hiển thị khi thiết bị ở chế độ dọc",
   "1. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị ở chế độ ngang",
   "Đăng nhập thành công",
   "1. Kiểm tra hiển thị khi thiết bị ở chế độ ngang",
   "1. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi chuyển đổi giữa chế độ dọc và ngang",
   "Đăng nhập thành công",
   "1. Kiểm tra hiển thị khi chuyển từ chế độ dọc sang ngang\n2. Kiểm tra hiển thị khi chuyển từ chế độ ngang sang dọc",
   "1 & 2. Không có lỗi gì xảy ra, giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ lớn nhất",
   "Đăng nhập thành công",
   "1. Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ lớn nhất",
   "1. Giao diện không bị vỡ")

tc("Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ nhỏ nhất",
   "Đăng nhập thành công",
   "1. Kiểm tra hiển thị khi thiết bị sử dụng cỡ chữ nhỏ nhất",
   "1. Giao diện không bị vỡ")

sub_label("Kiểm tra tương tác cơ bản với thiết bị")

tc("Xác nhận hiển thị khi chạm nút [Quay lại] trên Android",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Mở ứng dụng\n2. Chạm vào nút [Quay lại] trên thiết bị Android\n3. Xác nhận hiển thị",
   "3. Quay lại màn hình trước đó")

tc("Xác nhận hiển thị khi vuốt trái sang phải trên iOS",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Mở ứng dụng\n2. Vuốt từ trái sang phải trên thiết bị iOS\n3. Xác nhận hiển thị",
   "3. Quay lại màn hình trước đó")

tc("Xác nhận hiển thị khi tắt và mở lại ứng dụng",
   "Đăng nhập thành công",
   "1. Mở ứng dụng\n2. Tắt ứng dụng\n3. Mở lại ứng dụng\n4. Xác nhận hiển thị",
   "4. Ứng dụng mở lại từ Trang chủ, giữ session")

tc("Kiểm tra chế độ đa nhiệm (multitasking)",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Mở ứng dụng\n2. Trở về màn hình chính (không tắt ứng dụng)\n3. Mở lại ứng dụng\n4. Xác nhận hiển thị",
   "4. Ứng dụng giữ nguyên ở trạng thái hiện tại")

tc("Xác nhận hiển thị khi khóa và mở khóa màn hình",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Mở ứng dụng\n2. Khóa thiết bị\n3. Mở khóa thiết bị\n4. Xác nhận hiển thị",
   "4. Giữ nguyên trạng thái hiện tại của ứng dụng")

tc("Kiểm tra hành động kéo xuống để làm mới (pull-to-refresh)",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Người dùng ở màn hình Trang chủ\n2. Kéo xuống để làm mới",
   "2. Hiển thị dữ liệu mới nhất của màn hình",
   "CMR-13")

tc("Kiểm tra cuộn xuống để tải thêm (scroll-down-to-load-more)",
   "Đăng nhập thành công, danh sách có > 20 bản ghi",
   "1. Cuộn xuống cuối danh sách",
   "1. Hiển thị thêm dữ liệu mới",
   "CMR-04")

tc("Kiểm tra phản hồi khi nhận thông báo từ ứng dụng khác",
   "Đăng nhập thành công, đang ở Trang chủ",
   "1. Mở ứng dụng\n2. Ứng dụng khác gửi thông báo\n3. Xác nhận hiển thị",
   "3. Không có lỗi nào xảy ra")

# Save
wb.save(output_path)
print(f"Done! Total TCs: {tc_counter[0]}")
print(f"Saved to: {output_path}")
